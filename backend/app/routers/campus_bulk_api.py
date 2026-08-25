"""
Campus Bulk Upload — batch invite flow for freshers / campus drives.

Endpoints (all scoped to TA + Admin roles except two public session endpoints):
  POST /api/campus/upload                      — parse Excel, create batch + candidates
  GET  /api/campus/batch/{batch_id}            — paginated candidate list
  POST /api/campus/batch/{batch_id}/invite     — bulk invite selected candidates
  GET  /api/campus/batches                     — list batches for a requisition
  POST /api/campus/session/{token}/resume      — PUBLIC: candidate resume upload during NexAI
  GET  /api/campus/session/{token}/is-campus   — PUBLIC: does this token belong to a campus batch?
"""
import io
import json
import os
import re
import secrets

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, Query, UploadFile
from pydantic import BaseModel

from datetime import datetime

from ..db import query, query_one
from ..auth_utils import get_current_user
from ..services.resume_parser import extract_text as _parse_resume
from ..services import pipeline as _pipeline_svc
from ..services import prerender as _prerender_svc
from ..services import excel_export
from ..services.activity_log import log_activity

router = APIRouter(prefix="/api/campus", tags=["campus"])

# ── Column name auto-detection ────────────────────────────────────────────────
#
# College excels use wildly different header text ("Name of the Student",
# "EmailAddress", "MobileNo", "CPI" instead of CGPA, ...) and column order,
# and no two colleges send the same template. Instead of matching exact
# header strings, we normalise each header (strip spaces/punctuation, lower-
# case) and look for characteristic keyword substrings, most-specific first,
# so "EmailAddress" / "E-Mail" / "Mail Id" all resolve to "email" the same
# way. Fields are resolved in a fixed priority order so that e.g. "College
# Name" is claimed by `college` before the generic `name` pass ever sees it.

CANONICAL_FIELDS = [
    "name", "email", "phone", "college", "branch",
    "cgpa", "graduation_year", "current_company",
]

# (canonical, keywords-most-specific-first, exclude-substrings)
# Matching is done on the "squashed" header (lowercased, punctuation/spaces
# stripped) — a keyword match is `keyword in squashed_header`.
_FIELD_RULES: list[tuple[str, list[str], list[str]]] = [
    ("email", ["email", "mail"], []),
    ("phone", ["mobileno", "mobile", "phoneno", "phone",
               "contactno", "contact", "cellno", "cell", "whatsapp"], []),
    ("college", ["collegename", "college", "universityname", "university",
                 "institutename", "institute", "institution"], []),
    ("branch", ["branchname", "branch", "department", "dept", "stream",
                "specialisation", "specialization", "discipline", "course"], []),
    ("current_company", ["currentcompany", "currentemployer", "employer",
                          "organisation", "organization", "companyname",
                          "company"], []),
    # cgpa: strict keywords always win; generic "percentage/marks/score" only
    # count if the header isn't obviously a 10th/12th/semester-wise mark sheet
    # column (those are common alongside a genuine overall CGPA/CPI column).
    ("cgpa", ["cgpa", "cpi", "sgpa", "gpa", "aggregatepercentage",
              "aggregate", "overallpercentage", "overall"], []),
    ("cgpa", ["percentage", "percent", "marks", "score"],
     ["10th", "12th", "ssc", "hsc", "diploma", "semester", "sem1", "sem2",
      "sem3", "sem4", "sem5", "sem6", "sem7", "sem8", "backlog"]),
    ("graduation_year", ["graduationyear", "passoutyear", "passoutdate",
                          "yearofpassing", "yearofgraduation", "batchyear",
                          "passingyear"], []),
    ("graduation_year", ["year"],
     ["10th", "12th", "ssc", "hsc", "diploma", "semester", "admission",
      "birth", "dob", "enrollment", "enrolment", "rollno"]),
]

# Handled last so specific fields (college/company/etc.) claim their
# columns first — otherwise "College Name" / "Company Name" / "Father Name"
# would all get grabbed by the generic "name" pass.
_NAME_KEYWORDS = ["fullname", "candidatename", "studentname",
                  "nameofthestudent", "name"]
_NAME_EXCLUDE = ["college", "university", "institute", "company", "employer",
                  "father", "mother", "guardian", "parent", "spouse",
                  "husband", "wife", "permanent", "residence", "address",
                  "school", "branch", "department"]


def _squash(h: str) -> str:
    h = (h or "").lower().strip()
    h = re.sub(r"[^a-z0-9]+", "", h)
    return h


def _normalise_headers(
    raw_headers: list[str],
    overrides: dict[str, str] | None = None,
) -> dict[str, str]:
    """
    Return {canonical_field: original_header} for recognised columns.

    `overrides` (canonical_field -> raw header text, as sent by the frontend's
    manual "fix column mapping" step) takes priority over auto-detection for
    those specific fields; any remaining canonical fields are auto-detected
    from whatever headers are left.
    """
    result: dict[str, str] = {}
    used: set[int] = set()

    if overrides:
        for canonical, header_name in overrides.items():
            if not header_name:
                continue
            try:
                idx = raw_headers.index(header_name)
            except ValueError:
                continue
            if canonical in CANONICAL_FIELDS and idx not in used:
                result[canonical] = raw_headers[idx]
                used.add(idx)

    squashed = [_squash(h) for h in raw_headers]

    for canonical, keywords, exclude in _FIELD_RULES:
        if canonical in result:
            continue
        for kw in keywords:
            found = None
            for i, sq in enumerate(squashed):
                if i in used or not sq:
                    continue
                if kw in sq and not any(ex in sq for ex in exclude):
                    found = i
                    break
            if found is not None:
                result[canonical] = raw_headers[found]
                used.add(found)
                break

    if "name" not in result:
        for kw in _NAME_KEYWORDS:
            found = None
            for i, sq in enumerate(squashed):
                if i in used or not sq:
                    continue
                if kw in sq and not any(ex in sq for ex in _NAME_EXCLUDE):
                    found = i
                    break
            if found is not None:
                result["name"] = raw_headers[found]
                used.add(found)
                break

    return result


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _sniff_by_content(
    raw_headers: list[str],
    data_rows: list[tuple],
    mapping: dict[str, str],
) -> None:
    """
    Fallback for when header text gives no usable signal at all (e.g. a
    generic "Details"/"Info" column). Only applied to email/phone, whose
    value shapes are unambiguous — CGPA/graduation-year are left to header
    matching + manual override since numeric ranges alone are too easy to
    confuse with unrelated numeric columns (10th %, enrollment no, etc.).
    """
    used_headers = set(mapping.values())
    sample = [r for r in data_rows[:25] if r]

    for canonical, is_match in (
        ("email", lambda v: bool(_EMAIL_RE.match(v))),
        ("phone", lambda v: v.replace("+", "").isdigit() and 7 <= len(v.replace("+", "")) <= 13),
    ):
        if canonical in mapping:
            continue
        best_idx, best_hits = None, 0
        for i, h in enumerate(raw_headers):
            if not h or h in used_headers:
                continue
            hits = 0
            checked = 0
            for row in sample:
                if i >= len(row) or row[i] is None:
                    continue
                v = str(row[i]).strip()
                if not v:
                    continue
                checked += 1
                if is_match(v):
                    hits += 1
            if checked >= 2 and hits == checked and hits > best_hits:
                best_idx, best_hits = i, hits
        if best_idx is not None:
            mapping[canonical] = raw_headers[best_idx]
            used_headers.add(raw_headers[best_idx])


def _valid_email(v) -> bool:
    """
    Format-valid AND not a placeholder/example domain. Real institutional
    domains (college .ac.in, .edu, etc.) are unaffected — only the specific
    known-fake domains in email_validation._BLOCKED_DOMAINS are rejected.
    """
    if not v:
        return False
    from ..services.email_validation import assert_real_email
    try:
        assert_real_email(str(v).strip())
        return True
    except ValueError:
        return False


def _cell(row, headers: list[str], col_name: str | None):
    """Safe column read."""
    if col_name is None:
        return None
    try:
        idx = headers.index(col_name)
        v = row[idx]
        return str(v).strip() if v is not None else None
    except (ValueError, IndexError):
        return None


# ── Upload ────────────────────────────────────────────────────────────────────

@router.post("/upload", status_code=201)
async def upload_excel(
    file: UploadFile = File(...),
    requisition_id: str = Query(...),
    column_map: str | None = Form(None),
    user: dict = Depends(get_current_user),
):
    """
    Parse an .xlsx/.xls file and create a campus_upload_batch with campus_candidate rows.

    Column headers vary by college (different names, different order, and
    sometimes a title/banner row above the real header row) — see the
    auto-detection logic above `_normalise_headers`. `column_map` is an
    optional JSON object ({"cgpa": "CPI", ...}) from the frontend's manual
    "fix column mapping" step, used to correct any field auto-detection
    got wrong before the batch is committed.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")

    data = await file.read()
    if len(data) > 10 * 1024 * 1024:
        raise HTTPException(400, "File too large — maximum 10 MB")

    overrides: dict[str, str] = {}
    if column_map:
        try:
            parsed = json.loads(column_map)
            if isinstance(parsed, dict):
                overrides = {str(k): str(v) for k, v in parsed.items() if v}
        except (json.JSONDecodeError, TypeError):
            raise HTTPException(400, "Invalid column_map — must be a JSON object")

    if not query_one("SELECT id FROM requisition WHERE id=%s", [requisition_id]):
        raise HTTPException(404, "Requisition not found")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        raw_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        raise HTTPException(400, f"Cannot parse Excel file: {exc}")

    if not raw_rows:
        raise HTTPException(400, "Excel file is empty")

    # Some colleges add a title/banner row above the real header row — scan
    # the first few rows and use whichever one auto-detects the most fields.
    header_row_idx = 0
    raw_headers = [str(h) if h is not None else "" for h in raw_rows[0]]
    mapping = _normalise_headers(raw_headers)
    best_score = len(mapping)
    for ridx in range(1, min(5, len(raw_rows))):
        candidate_headers = [str(h) if h is not None else "" for h in raw_rows[ridx]]
        if not any(candidate_headers):
            continue
        candidate_mapping = _normalise_headers(candidate_headers)
        if len(candidate_mapping) > best_score:
            best_score = len(candidate_mapping)
            header_row_idx = ridx
            raw_headers = candidate_headers
            mapping = candidate_mapping

    data_rows = raw_rows[header_row_idx + 1:]

    if overrides:
        mapping = _normalise_headers(raw_headers, overrides=overrides)
    _sniff_by_content(raw_headers, data_rows, mapping)

    canonical_header_set = set(mapping.values())

    # Indices of unrecognised columns → go into extra_data
    extra_indices = [
        i for i, h in enumerate(raw_headers)
        if h and h not in canonical_header_set
    ]

    total_rows = 0
    skipped_rows = 0
    skipped_row_details: list[dict] = []
    valid_candidates: list[dict] = []

    for row_num, raw_row in enumerate(data_rows, start=header_row_idx + 2):
        # Skip blank rows
        if all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        total_rows += 1

        email_raw = _cell(raw_row, raw_headers, mapping.get("email"))
        if not _valid_email(email_raw):
            skipped_rows += 1
            # Previously just a count -- a recruiter had no way to find and
            # fix these rows, unlike the header-mapping mismatch flow, which
            # already gets a full manual-fix UI.
            skipped_row_details.append({
                "row": row_num,
                "name": _cell(raw_row, raw_headers, mapping.get("name")) or "",
                "email_raw": email_raw or "",
                "reason": "missing or invalid email",
            })
            continue

        # CGPA
        cgpa = None
        cgpa_raw = _cell(raw_row, raw_headers, mapping.get("cgpa"))
        if cgpa_raw:
            try:
                cgpa = round(float(cgpa_raw.replace("%", "")), 2)
            except (ValueError, AttributeError):
                pass

        # Graduation year
        grad_year = None
        gy_raw = _cell(raw_row, raw_headers, mapping.get("graduation_year"))
        if gy_raw:
            try:
                grad_year = int(float(gy_raw))
            except (ValueError, TypeError):
                pass

        # Extra columns
        extra_data: dict = {}
        for i in extra_indices:
            if i < len(raw_row) and raw_row[i] is not None:
                extra_data[raw_headers[i]] = str(raw_row[i]).strip()

        valid_candidates.append({
            "name":            _cell(raw_row, raw_headers, mapping.get("name")),
            "email":           email_raw.strip().lower(),
            "phone":           _cell(raw_row, raw_headers, mapping.get("phone")),
            "college":         _cell(raw_row, raw_headers, mapping.get("college")),
            "branch":          _cell(raw_row, raw_headers, mapping.get("branch")),
            "cgpa":            cgpa,
            "graduation_year": grad_year,
            "current_company": _cell(raw_row, raw_headers, mapping.get("current_company")),
            "extra_data":      extra_data,
        })

    if not valid_candidates:
        raise HTTPException(
            400,
            "No valid candidates found — all rows have blank or invalid email addresses",
        )

    # Persist batch
    batch_row = query_one(
        """INSERT INTO campus_upload_batch
               (requisition_id, uploaded_by, file_name, total_rows)
           VALUES (%s, %s, %s, %s) RETURNING id""",
        [requisition_id, user["sub"], file.filename, total_rows],
    )
    batch_id = str(batch_row["id"])

    for c in valid_candidates:
        inserted = query_one(
            """INSERT INTO campus_candidate
                   (batch_id, requisition_id, name, email, phone,
                    college, branch, cgpa, graduation_year, current_company, extra_data)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
               RETURNING id""",
            [
                batch_id, requisition_id,
                c["name"], c["email"], c["phone"],
                c["college"], c["branch"], c["cgpa"],
                c["graduation_year"], c["current_company"], json.dumps(c["extra_data"]),
            ],
        )
        c["id"] = str(inserted["id"])

    detected = {k: (k in mapping) for k in CANONICAL_FIELDS}
    preview = valid_candidates[:10]

    log_activity(
        "campus_batch", "campus_upload",
        entity_id=batch_id, requisition_id=requisition_id,
        actor_id=user["sub"], actor_role=user["role"],
        detail={"file_name": file.filename, "total_rows": total_rows,
                "valid_rows": len(valid_candidates), "skipped_rows": skipped_rows},
    )

    return {
        "batch_id":    batch_id,
        "total_rows":  total_rows,
        "valid_rows":  len(valid_candidates),
        "skipped_rows": skipped_rows,
        "skipped_row_details": skipped_row_details,
        "detected":    detected,
        "raw_headers": [h for h in raw_headers if h],
        "column_map":  mapping,
        "preview":     preview,
    }


# ── Batch detail (paginated) ──────────────────────────────────────────────────

# Joins in the real interview outcome (nexai_session / application) alongside
# the campus_candidate staging row, so Candidate Preview can show what actually
# happened to an invited candidate — not just that an invite was queued.
# `invite_status` on campus_candidate only ever reaches 'invite_queued'/'invited'
# in practice (nothing sets it to 'interview_started'/'completed' even though
# the CHECK constraint allows those values) — the real progress lives on
# nexai_session.status once the candidate's application is linked.
_CANDIDATE_JOIN_SELECT_BASE = """
    SELECT cc.id, cc.name, cc.email, cc.phone, cc.college, cc.branch,
           cc.cgpa, cc.graduation_year, cc.current_company,
           cc.invite_status, cc.invite_sent_at, cc.resume_uploaded, cc.resume_url,
           cc.nexai_session_id, cc.created_at, cc.batch_id,
           cub.file_name AS batch_file_name,
           ns.status AS session_status, ns.completed_at AS interview_completed_at,
           a.combined_score, a.bot_score, ns.raw_score
    FROM campus_candidate cc
    JOIN campus_upload_batch cub ON cub.id = cc.batch_id
    LEFT JOIN application   a  ON a.id = cc.application_id
    LEFT JOIN nexai_session ns ON ns.application_id = cc.application_id
    WHERE {where}
    ORDER BY cc.created_at, cc.id
"""
_CANDIDATE_JOIN_SELECT = _CANDIDATE_JOIN_SELECT_BASE.format(where="cc.batch_id = %s")
# Aggregate, cross-batch view: every candidate ever sent for a requisition,
# not just the most recently uploaded batch — see get_requisition_candidates.
_CANDIDATE_JOIN_SELECT_BY_REQ = _CANDIDATE_JOIN_SELECT_BASE.format(where="cc.requisition_id = %s")


def _candidate_score(c: dict) -> float | None:
    for v in (c.get("combined_score"), c.get("bot_score"), c.get("raw_score")):
        if v is not None:
            return float(v)
    return None


def _candidate_status(c: dict) -> str:
    session_status = c.get("session_status")
    if session_status == "completed":
        return "completed"
    if session_status == "in_progress":
        return "interview_started"
    if session_status == "failed":
        return "interview_failed"
    return c["invite_status"]


def _serialize_candidate(c: dict) -> dict:
    return {
        "id":                str(c["id"]),
        "batch_id":          str(c["batch_id"]),
        "batch_file_name":   c.get("batch_file_name"),
        "name":              c["name"],
        "email":             c["email"],
        "phone":             c["phone"],
        "college":           c["college"],
        "branch":            c["branch"],
        "cgpa":              float(c["cgpa"]) if c["cgpa"] is not None else None,
        "graduation_year":   c["graduation_year"],
        "current_company":   c["current_company"],
        "invite_status":     c["invite_status"],
        "status":            _candidate_status(c),
        "score":             _candidate_score(c),
        "invite_sent_at":    c["invite_sent_at"].isoformat() if c["invite_sent_at"] else None,
        "interview_completed_at": c["interview_completed_at"].isoformat() if c["interview_completed_at"] else None,
        "resume_uploaded":   c["resume_uploaded"],
        "resume_url":        c["resume_url"],
        "nexai_link":        (
            c["nexai_session_id"]
            if c["invite_status"] == "invite_queued" and c["nexai_session_id"]
            else None
        ),
    }


@router.get("/batch/{batch_id}")
def get_batch(
    batch_id: str,
    page: int = Query(1, ge=1),
    user: dict = Depends(get_current_user),
):
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    batch = query_one("SELECT * FROM campus_upload_batch WHERE id=%s", [batch_id])
    if not batch:
        raise HTTPException(404, "Batch not found")

    per_page = 50
    offset = (page - 1) * per_page
    candidates = query(
        _CANDIDATE_JOIN_SELECT + " LIMIT %s OFFSET %s",
        [batch_id, per_page, offset],
    )
    total = (query_one(
        "SELECT COUNT(*) AS n FROM campus_candidate WHERE batch_id=%s", [batch_id]
    ) or {}).get("n", 0)

    return {
        "batch": {
            "id":             str(batch["id"]),
            "requisition_id": str(batch["requisition_id"]),
            "file_name":      batch["file_name"],
            "total_rows":     batch["total_rows"],
            "selected_count": batch["selected_count"],
            "invited_count":  batch["invited_count"],
            "status":         batch["status"],
            "created_at":     batch["created_at"].isoformat() if batch["created_at"] else None,
        },
        "candidates": [_serialize_candidate(c) for c in candidates],
        "total": total,
        "page":  page,
        "pages": max(1, -(-total // per_page)),
    }


# ── Requisition-wide candidate list (all batches, aggregate view) ────────────

@router.get("/requisition/{requisition_id}/candidates")
def get_requisition_candidates(
    requisition_id: str,
    page: int = Query(1, ge=1),
    user: dict = Depends(get_current_user),
):
    """
    Every campus candidate ever sent for this requisition, across every
    upload batch — not just the most recently uploaded one. This is the
    running "who has ever been invited/interviewed for this requisition"
    list a recruiter needs when campus drives arrive in multiple Excel
    batches over time. `batch_file_name` on each row identifies which
    upload a candidate originally came from.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    if not query_one("SELECT id FROM requisition WHERE id=%s", [requisition_id]):
        raise HTTPException(404, "Requisition not found")

    per_page = 50
    offset = (page - 1) * per_page
    candidates = query(
        _CANDIDATE_JOIN_SELECT_BY_REQ + " LIMIT %s OFFSET %s",
        [requisition_id, per_page, offset],
    )
    total = (query_one(
        "SELECT COUNT(*) AS n FROM campus_candidate WHERE requisition_id=%s", [requisition_id]
    ) or {}).get("n", 0)

    return {
        "requisition_id": requisition_id,
        "candidates": [_serialize_candidate(c) for c in candidates],
        "total": total,
        "page":  page,
        "pages": max(1, -(-total // per_page)),
    }


@router.get("/requisition/{requisition_id}/export")
def export_requisition_candidates(requisition_id: str, user: dict = Depends(get_current_user)):
    """
    Export every campus candidate ever sent for this requisition (across all
    upload batches) to .xlsx — the aggregate counterpart of /batch/{id}/export.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    req = query_one("SELECT title, req_code FROM requisition WHERE id=%s", [requisition_id])
    if not req:
        raise HTTPException(404, "Requisition not found")

    candidates = query(_CANDIDATE_JOIN_SELECT_BY_REQ, [requisition_id])

    status_labels = {
        "pending": "Pending", "invite_queued": "Invite Queued", "invited": "Invited",
        "interview_started": "Interview In Progress", "completed": "Interview Completed",
        "interview_failed": "Interview Failed",
    }

    rows = [
        {
            "Name":               c["name"] or "",
            "Email":              c["email"] or "",
            "Phone":              c["phone"] or "",
            "College":            c["college"] or "",
            "Branch":             c["branch"] or "",
            "CGPA":               float(c["cgpa"]) if c["cgpa"] is not None else None,
            "Graduation Year":    c["graduation_year"],
            "Current Company":    c["current_company"] or "",
            "Status":             status_labels.get(_candidate_status(c), _candidate_status(c)),
            "Score":              _candidate_score(c),
            "CV Uploaded":        "Yes" if c["resume_uploaded"] else "No",
            "Upload Batch":       c["batch_file_name"] or "",
            "Invited At":         c["invite_sent_at"].isoformat() if c["invite_sent_at"] else "",
            "Interview Completed At": c["interview_completed_at"].isoformat() if c["interview_completed_at"] else "",
        }
        for c in candidates
    ]

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    excel_export.sheet_from_rows(wb, "Campus Pool", rows)
    excel_export.build_summary_sheet(
        wb,
        title=f"Campus Pool — {req.get('req_code') or ''} {req.get('title') or requisition_id}",
        generated_by=user.get("name") or user.get("email") or "",
        generated_at=datetime.now(),
        filters_applied=[],
        rows=rows,
        measures_meta=[{"key": "Score", "label": "Score"}, {"key": "CGPA", "label": "CGPA"}],
    )
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", req.get("req_code") or req.get("title") or requisition_id)
    return excel_export.stream_workbook(wb, f"enternly_campus_pool_{safe_name}.xlsx")


# ── Export whole campus pool to Excel ─────────────────────────────────────────

@router.get("/batch/{batch_id}/export")
def export_batch(batch_id: str, user: dict = Depends(get_current_user)):
    """
    Export every candidate in this batch (not just the current page) to an
    .xlsx, including real interview outcome/score and CV upload status —
    the same enriched fields shown in Candidate Preview.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    batch = query_one("SELECT * FROM campus_upload_batch WHERE id=%s", [batch_id])
    if not batch:
        raise HTTPException(404, "Batch not found")

    candidates = query(_CANDIDATE_JOIN_SELECT, [batch_id])

    status_labels = {
        "pending": "Pending", "invite_queued": "Invite Queued", "invited": "Invited",
        "interview_started": "Interview In Progress", "completed": "Interview Completed",
        "interview_failed": "Interview Failed",
    }

    rows = [
        {
            "Name":               c["name"] or "",
            "Email":              c["email"] or "",
            "Phone":              c["phone"] or "",
            "College":            c["college"] or "",
            "Branch":             c["branch"] or "",
            "CGPA":               float(c["cgpa"]) if c["cgpa"] is not None else None,
            "Graduation Year":    c["graduation_year"],
            "Current Company":    c["current_company"] or "",
            "Status":             status_labels.get(_candidate_status(c), _candidate_status(c)),
            "Score":              _candidate_score(c),
            "CV Uploaded":        "Yes" if c["resume_uploaded"] else "No",
            "Invited At":         c["invite_sent_at"].isoformat() if c["invite_sent_at"] else "",
            "Interview Completed At": c["interview_completed_at"].isoformat() if c["interview_completed_at"] else "",
        }
        for c in candidates
    ]

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    excel_export.sheet_from_rows(wb, "Campus Pool", rows)
    excel_export.build_summary_sheet(
        wb,
        title=f"Campus Pool — {batch['file_name'] or batch_id}",
        generated_by=user.get("name") or user.get("email") or "",
        generated_at=datetime.now(),
        filters_applied=[],
        rows=rows,
        measures_meta=[{"key": "Score", "label": "Score"}, {"key": "CGPA", "label": "CGPA"}],
    )
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", batch["file_name"] or "campus_pool")
    return excel_export.stream_workbook(wb, f"enternly_campus_pool_{safe_name}.xlsx")


# ── Bulk invite ───────────────────────────────────────────────────────────────

class BulkCampusInviteIn(BaseModel):
    candidate_ids: list[str]
    requisition_id: str


def _campus_base_url() -> tuple[str, bool]:
    """
    Return (base_url, is_localhost).
    Uses connectors._load_email_cfg() which already applies the correct
    priority: DB (Settings UI) → APP_BASE_URL env var → default localhost.
    DB always wins, so the Settings UI value is never shadowed by .env.prod.
    """
    from ..services.connectors import _load_email_cfg
    url = (_load_email_cfg().get("base_url") or "http://localhost:8000").strip().rstrip("/")
    is_local = any(x in url for x in ("localhost", "127.0.0.1", "0.0.0.0"))
    return url, is_local


@router.post("/batch/{batch_id}/invite")
def bulk_invite(
    batch_id: str,
    body: BulkCampusInviteIn,
    background_tasks: BackgroundTasks,
    user: dict = Depends(get_current_user),
):
    """
    For each candidate_id:
      1. Upsert candidate + application records.
      2. Generate NexAI invite token.
      3. Queue the invite email — actual sending happens off-request in
         campus_email_worker.py, which sends in throttled batches of 20
         so a 1000+ row campus drive doesn't fire all its emails at once.
      4. Update batch counts.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    batch = query_one("SELECT * FROM campus_upload_batch WHERE id=%s", [batch_id])
    if not batch:
        raise HTTPException(404, "Batch not found")

    req = query_one(
        """SELECT r.id, r.title, r.key_skills, r.job_description,
                  gc.name AS company
           FROM requisition r
           JOIN business_unit bu ON bu.id = r.bu_id
           JOIN group_company gc ON gc.id = bu.company_id
           WHERE r.id=%s""",
        [body.requisition_id],
    )
    if not req:
        raise HTTPException(404, "Requisition not found")

    base_url, is_local = _campus_base_url()

    queued_for_email = 0
    queued_local = 0
    failed: list[dict] = []

    # Lazy import to avoid circular dependency
    from ..routers.nexai_api import _generate_questions
    from ..services.pipeline import _check_no_poach_block, NoPoachBlockedError

    blocked: list[dict] = []

    for cid in body.candidate_ids:
        campus_c = query_one(
            "SELECT * FROM campus_candidate WHERE id=%s AND batch_id=%s",
            [cid, batch_id],
        )
        if not campus_c:
            failed.append({"id": cid, "reason": "not_found"})
            continue

        # No-poach hard block, run BEFORE candidate/application creation --
        # this is the actual "add candidate" moment for campus intake (the
        # candidate/application rows are created fresh in this loop, unlike
        # the resume-attach step elsewhere, which runs after they already
        # exist and could only flag, not block). Matches the same gate
        # already enforced on career-site/vendor intake.
        try:
            _check_no_poach_block(campus_c.get("current_company"), body.requisition_id, entity_id=cid)
        except NoPoachBlockedError as exc:
            # Migration 60 added 'blocked' to invite_status's CHECK constraint
            # so this is now queryable on the row itself, not just the
            # response/activity_log. The candidate/application still never
            # gets created (matches career-site/vendor/portal: retain the
            # source row, never silently drop it, but don't let it into the
            # pipeline).
            query(
                "UPDATE campus_candidate SET invite_status='blocked' WHERE id=%s",
                [cid], fetch=False,
            )
            blocked.append({
                "id": cid, "name": campus_c.get("name"), "email": campus_c.get("email"),
                "current_company": campus_c.get("current_company"),
                "matched_company": exc.company_name,
            })
            continue

        # 1 — upsert candidate
        existing_cand = query_one(
            "SELECT id FROM candidate WHERE LOWER(email)=LOWER(%s)",
            [campus_c["email"]],
        )
        if existing_cand:
            cand_id = str(existing_cand["id"])
        else:
            new_cand = query_one(
                """INSERT INTO candidate (full_name, email, phone, source)
                   VALUES (%s, %s, %s, 'campus_bulk')
                   ON CONFLICT DO NOTHING RETURNING id""",
                [campus_c["name"] or "Unknown", campus_c["email"], campus_c["phone"]],
            )
            if not new_cand:
                # Race: inserted by another request between our check and insert
                new_cand = query_one(
                    "SELECT id FROM candidate WHERE LOWER(email)=LOWER(%s)",
                    [campus_c["email"]],
                )
            if not new_cand:
                failed.append({"id": cid, "reason": "candidate_upsert_failed"})
                continue
            cand_id = str(new_cand["id"])

        # 2 — upsert application
        existing_app = query_one(
            "SELECT id FROM application WHERE candidate_id=%s AND requisition_id=%s",
            [cand_id, body.requisition_id],
        )
        if existing_app:
            app_id = str(existing_app["id"])
        else:
            new_app = query_one(
                """INSERT INTO application
                       (candidate_id, requisition_id, status, applied_at)
                   VALUES (%s, %s, 'nexai_bot', now()) RETURNING id""",
                [cand_id, body.requisition_id],
            )
            if not new_app:
                failed.append({"id": cid, "reason": "application_create_failed"})
                continue
            app_id = str(new_app["id"])

        # 3 — generate invite token (skip if active invite already exists)
        active = query_one(
            """SELECT id FROM nexai_invite
               WHERE application_id=%s AND used_at IS NULL AND expires_at > now()
               LIMIT 1""",
            [app_id],
        )
        if active:
            # Re-use existing token for the link
            token_row = query_one(
                "SELECT token FROM nexai_invite WHERE id=%s", [active["id"]]
            )
            token = token_row["token"] if token_row else secrets.token_urlsafe(32)
        else:
            token = secrets.token_urlsafe(32)
            try:
                query(
                    """INSERT INTO nexai_invite (application_id, token, created_by)
                       VALUES (%s, %s, %s)""",
                    [app_id, token, user["sub"]],
                    fetch=False,
                )
            except Exception as exc:
                failed.append({"id": cid, "reason": f"token_error: {exc}"})
                continue

        # 4 — upsert nexai_session for avatar pre-render
        _saved_qs = query_one(
            "SELECT questions FROM requisition_questions WHERE requisition_id=%s",
            [body.requisition_id],
        )
        questions = (
            list(_saved_qs["questions"]) if _saved_qs
            else _generate_questions(
                req.get("key_skills") or [],
                req.get("job_description") or "",
            )
        )
        existing_sess = query_one(
            "SELECT id FROM nexai_session WHERE application_id=%s", [app_id]
        )
        if existing_sess:
            session_id = str(existing_sess["id"])
        else:
            sess_row = query_one(
                """INSERT INTO nexai_session
                       (application_id, requisition_id, questions, status)
                   VALUES (%s, %s, %s::jsonb, 'pending') RETURNING id""",
                [app_id, body.requisition_id, json.dumps(questions)],
            )
            session_id = str(sess_row["id"])

        background_tasks.add_task(_prerender_svc.prerender_interview_videos, session_id)

        invite_url = f"{base_url}/nexai-interview?token={token}"

        # 5 — update campus_candidate with application_id + link
        query(
            """UPDATE campus_candidate
               SET application_id=%s, nexai_session_id=%s, invite_sent_at=now()
               WHERE id=%s""",
            [app_id, invite_url if is_local else token, cid],
            fetch=False,
        )

        if is_local:
            # PROD_BASE_URL not set or is localhost — queue, do not email
            query(
                "UPDATE campus_candidate SET invite_status='invite_queued' WHERE id=%s",
                [cid], fetch=False,
            )
            queued_local += 1
        else:
            # Hand off to campus_email_worker.py — it sends in throttled
            # batches of 20 instead of blocking this request for 1000+ rows.
            query(
                """UPDATE campus_candidate
                   SET invite_status='invite_queued', email_status='queued',
                       email_attempts=0, email_error=NULL, email_next_attempt_at=NULL
                   WHERE id=%s""",
                [cid], fetch=False,
            )
            queued_for_email += 1

    # Update batch counters
    total_actioned = len(body.candidate_ids) - len(failed) - len(blocked)
    total_sent = queued_for_email + queued_local
    if total_actioned > 0:
        query(
            """UPDATE campus_upload_batch
               SET selected_count = selected_count + %s,
                   invited_count  = invited_count  + %s,
                   status = CASE WHEN status='draft' THEN 'invites_sent' ELSE status END
               WHERE id=%s""",
            [total_actioned, total_sent, batch_id],
            fetch=False,
        )

    log_activity(
        "campus_batch", "campus_invite_sent",
        entity_id=batch_id, requisition_id=body.requisition_id,
        actor_id=user["sub"], actor_role=user["role"],
        detail={"requested": len(body.candidate_ids), "queued_for_email": queued_for_email,
                "queued_local": queued_local, "failed_count": len(failed),
                "blocked_count": len(blocked)},
    )
    if blocked:
        log_activity(
            "campus_batch", "campus_no_poach_blocked",
            entity_id=batch_id, requisition_id=body.requisition_id,
            actor_id=user["sub"], actor_role=user["role"],
            detail={"blocked": blocked},
        )

    return {
        "queued_for_email": queued_for_email,
        "queued_local":      queued_local,
        "failed":            failed,
        "blocked":           blocked,
    }


# ── Resend queued invites ─────────────────────────────────────────────────────

@router.post("/batch/{batch_id}/resend-queued")
def resend_queued_invites(
    batch_id: str,
    user: dict = Depends(get_current_user),
):
    """
    Re-queue candidates whose invite email permanently failed (email_status='failed',
    e.g. after fixing SMTP creds in Settings) so campus_email_worker.py picks them up
    again in its next throttled batch. Does not send anything synchronously.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    base_url, is_local = _campus_base_url()
    if is_local:
        raise HTTPException(400, "Production URL is still localhost. Set the base URL in Settings first.")

    batch = query_one("SELECT * FROM campus_upload_batch WHERE id=%s", [batch_id])
    if not batch:
        raise HTTPException(404, "Batch not found")

    failed_rows = query(
        """SELECT id FROM campus_candidate
           WHERE batch_id=%s AND invite_status='invite_queued' AND email_status='failed'""",
        [batch_id],
    )
    if not failed_rows:
        return {"requeued": 0}

    ids = [str(r["id"]) for r in failed_rows]
    query(
        """UPDATE campus_candidate
           SET email_status='queued', email_attempts=0,
               email_error=NULL, email_next_attempt_at=NULL
           WHERE id = ANY(%s::uuid[])""",
        [ids],
        fetch=False,
    )

    return {"requeued": len(ids)}


# ── Batch list ────────────────────────────────────────────────────────────────

@router.get("/batches")
def list_batches(
    requisition_id: str = Query(...),
    user: dict = Depends(get_current_user),
):
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    rows = query(
        """SELECT cub.id, cub.file_name, cub.total_rows, cub.selected_count,
                  cub.invited_count, cub.status, cub.created_at
           FROM campus_upload_batch cub
           JOIN requisition r ON r.id = cub.requisition_id
           WHERE cub.requisition_id=%s AND r.tenant_id=%s
           ORDER BY cub.created_at DESC""",
        [requisition_id, user.get("tenant_id")],
    )
    return [
        {
            "id":             str(r["id"]),
            "file_name":      r["file_name"],
            "total_rows":     r["total_rows"],
            "selected_count": r["selected_count"],
            "invited_count":  r["invited_count"],
            "status":         r["status"],
            "created_at":     r["created_at"].isoformat() if r["created_at"] else None,
        }
        for r in rows
    ]


@router.delete("/batch/{batch_id}")
def delete_batch(batch_id: str, user: dict = Depends(get_current_user)):
    """
    Delete a wrongly-uploaded batch and its staged candidate rows.

    Only ever deletes campus_candidate staging rows + the batch itself — any
    application/candidate/nexai_invite already created for a candidate whose
    invite was sent is left untouched, so this can't accidentally remove a
    real candidate from the pipeline.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    batch = query_one(
        """SELECT cub.id FROM campus_upload_batch cub
           JOIN requisition r ON r.id = cub.requisition_id
           WHERE cub.id=%s AND r.tenant_id=%s""",
        [batch_id, user.get("tenant_id")],
    )
    if not batch:
        raise HTTPException(404, "Batch not found")

    query("DELETE FROM campus_candidate WHERE batch_id=%s", [batch_id], fetch=False)
    query("DELETE FROM campus_upload_batch WHERE id=%s", [batch_id], fetch=False)
    return {"ok": True}


# ── Delete a single candidate row ────────────────────────────────────────────

@router.delete("/candidate/{candidate_id}")
def delete_candidate(candidate_id: str, user: dict = Depends(get_current_user)):
    """
    Remove one wrongly-added candidate row (typo, wrong person, duplicate)
    from its batch, so a recruiter can fix that entry and re-add/re-upload
    it without discarding the rest of the batch.

    Same rule as batch delete: only the campus_candidate staging row is
    removed — any application/candidate/nexai_invite already created for a
    candidate whose invite was already sent is left untouched.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")

    cand = query_one("SELECT * FROM campus_candidate WHERE id=%s", [candidate_id])
    if not cand:
        raise HTTPException(404, "Candidate not found")

    query("DELETE FROM campus_candidate WHERE id=%s", [candidate_id], fetch=False)

    was_actioned = cand["invite_status"] != "pending"
    query(
        """UPDATE campus_upload_batch
           SET total_rows     = GREATEST(total_rows - 1, 0),
               selected_count = GREATEST(selected_count - %s, 0),
               invited_count  = GREATEST(invited_count  - %s, 0)
           WHERE id=%s""",
        [1 if was_actioned else 0, 1 if was_actioned else 0, cand["batch_id"]],
        fetch=False,
    )
    return {"ok": True}


# ── Public: campus session resume upload ──────────────────────────────────────

@router.post("/session/{session_token}/resume")
async def upload_campus_resume(
    session_token: str,
    file: UploadFile = File(...),
):
    """
    Public (no JWT) — candidate uploads resume during a campus NexAI session.
    Runs intake_and_screen with is_fresher_role forced True, updates campus_candidate.
    """
    invite = query_one(
        """SELECT ni.application_id, a.requisition_id
           FROM nexai_invite ni
           JOIN application a ON a.id = ni.application_id
           WHERE ni.token=%s AND ni.expires_at > now()""",
        [session_token],
    )
    if not invite:
        raise HTTPException(404, "Invalid or expired session token")

    app_id = str(invite["application_id"])
    req_id = str(invite["requisition_id"])

    if not file.filename or not file.filename.lower().endswith((".pdf", ".docx", ".doc")):
        raise HTTPException(400, "Only PDF or Word documents are accepted (PDF/DOCX/DOC)")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(400, "File too large — maximum 5 MB")

    _CV_STORE = os.environ.get("CV_STORE_DIR", "/app/cv_store")
    os.makedirs(_CV_STORE, exist_ok=True)
    import uuid as _uuid
    safe_name = f"campus_{_uuid.uuid4().hex}_{os.path.basename(file.filename).replace(' ', '_')}"
    file_path = os.path.join(_CV_STORE, safe_name)
    with open(file_path, "wb") as fh:
        fh.write(content)

    resume_url = f"/api/resume/{safe_name}"

    # Parse resume text and run screening
    try:
        resume_text = _parse_resume(file_path)
    except Exception:
        resume_text = ""

    cand_row = query_one(
        "SELECT candidate_id FROM application WHERE id=%s", [app_id]
    )
    campus_row = query_one(
        "SELECT current_company FROM campus_candidate WHERE application_id=%s", [app_id]
    )
    if cand_row and resume_text:
        try:
            _pipeline_svc.intake_and_screen(
                requisition_id=req_id,
                candidate_id=str(cand_row["candidate_id"]),
                resume_text=resume_text,
                candidate_years=0.0,
                file_size_bytes=len(content),
                current_company=(campus_row or {}).get("current_company"),
            )
            if campus_row and campus_row.get("current_company"):
                query(
                    "UPDATE application SET current_company=%s WHERE id=%s",
                    [campus_row["current_company"], app_id], fetch=False,
                )
        except _pipeline_svc.NoPoachBlockedError as exc:
            # NOTE: unlike career-site/vendor intake, the campus_candidate +
            # application rows already exist by this point in the bulk-upload
            # flow (this step only attaches/scores the resume afterward) --
            # so this can't be a true pre-creation block. Flag it distinctly
            # instead of letting it fall into the generic except below, so a
            # recruiter reviewing the batch sees "no-poach match" rather than
            # a swallowed screening failure.
            print(f"[campus-resume] no-poach blocked for application {app_id}: {exc}")
            query(
                """UPDATE application
                     SET flags = flags || jsonb_build_object('no_poach_blocked',
                           jsonb_build_object('matched_company', %s::text))
                   WHERE id = %s""",
                [exc.company_name, app_id], fetch=False,
            )
        except Exception as exc:
            print(f"[campus-resume] intake_and_screen failed: {exc}")

    # Update candidate record with resume_url
    if cand_row:
        query(
            "UPDATE candidate SET resume_url=%s WHERE id=%s",
            [resume_url, str(cand_row["candidate_id"])],
            fetch=False,
        )

        # Also ingest into the CV Repository so campus resumes are searchable
        # and AI-enriched like every other intake path (career site, vendor,
        # admin add-candidate). Never blocks the upload response on failure.
        try:
            from ..routers.cv_api import ingest_and_link
            ingest_and_link(
                data=content,
                filename=file.filename or f"campus_{app_id}.pdf",
                source="application",
                uploaded_by=None,
                candidate_id=str(cand_row["candidate_id"]),
                req_id=req_id,
            )
        except Exception as exc:
            print(f"[campus-resume] CV repository ingest failed for candidate "
                  f"{cand_row['candidate_id']}: {exc}")

    # Mark campus_candidate as uploaded
    query(
        """UPDATE campus_candidate
           SET resume_uploaded=TRUE, resume_url=%s
           WHERE application_id=%s""",
        [resume_url, app_id],
        fetch=False,
    )

    return {"ok": True, "resume_url": resume_url}


# ── Public: is-campus check ───────────────────────────────────────────────────

@router.get("/session/{session_token}/is-campus")
def is_campus_session(session_token: str):
    """Public (no JWT) — returns whether this invite belongs to a campus bulk batch."""
    invite = query_one(
        "SELECT application_id FROM nexai_invite WHERE token=%s",
        [session_token],
    )
    if not invite:
        return {"is_campus": False}

    campus_c = query_one(
        "SELECT id FROM campus_candidate WHERE application_id=%s LIMIT 1",
        [str(invite["application_id"])],
    )
    return {"is_campus": campus_c is not None}
