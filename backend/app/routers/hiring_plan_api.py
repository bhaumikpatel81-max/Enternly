"""
Hiring Plan — budget sheet import, manual entry, requisition linking,
demand-vs-fulfilled dashboard.

All endpoints are restricted to ta_manager, recruiter, and admin.
Computed fields (budgeted_total, months, prorata_budget, etc.) are calculated
in Python on every read/write — never stored.
"""
import io
import re
import uuid
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel

from ..db import query, query_one
from ..auth_utils import get_current_user
from ..services import excel_export

router = APIRouter(prefix="/api/hiring-plan", tags=["hiring-plan"])

_ALLOWED = {"ta_manager", "recruiter", "admin"}

HP_STATUSES = ('Open Position', 'Offered', 'Joined', 'Hold', 'Internal Employee')


def _require(user: dict):
    if user["role"] not in _ALLOWED:
        raise HTTPException(403, "Hiring Plan: ta_manager / recruiter / admin only")


# ── Computed-field helpers ────────────────────────────────────────────────────

def _to_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y',
                '%d %b %Y', '%b %d %Y', '%B %d, %Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None  # non-date text like 'Hold' — silently ignore


def _months_between(d1: date, d2: date) -> int:
    """Whole months from d1 to d2 — DATEDIF 'M' semantics."""
    if not d1 or not d2 or d2 <= d1:
        return 0
    total = (d2.year - d1.year) * 12 + (d2.month - d1.month)
    if d2.day < d1.day:
        total -= 1
    return max(0, total)


def _safe_float(v) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _safe_int(v, default: int = 1) -> int:
    try:
        return max(1, int(float(v or default)))
    except (TypeError, ValueError):
        return default


def _compute(row: dict) -> dict:
    """Return a new dict with all computed columns attached."""
    r = dict(row)

    bf  = _safe_float(r.get('budgeted_fixed'))
    bv  = _safe_float(r.get('budgeted_variable'))
    of_ = _safe_float(r.get('offered_fixed'))
    ov  = _safe_float(r.get('offered_variable'))
    status = r.get('hiring_status') or 'Open Position'

    budgeted_total = bf + bv
    offered_total  = of_ + ov

    months = 0
    if status not in ('Hold', 'Internal Employee'):
        pod = _to_date(r.get('planned_onboarding_date'))
        sbt = _to_date(r.get('salary_budgeted_till'))
        if pod and sbt:
            months = _months_between(pod, sbt)

    per_month_budget  = budgeted_total / 12 if budgeted_total else 0.0
    prorata_budget    = per_month_budget * months
    per_month_offered = offered_total / 12 if offered_total else 0.0
    prorata_offered   = per_month_offered * months
    actual_costing    = prorata_offered if prorata_offered > 0 else prorata_budget

    r['budgeted_total']      = round(budgeted_total, 2)
    r['months']              = months
    r['per_month_budget']    = round(per_month_budget, 2)
    r['prorata_budget']      = round(prorata_budget, 2)
    r['offered_total']       = round(offered_total, 2)
    r['budget_offered_diff'] = round(budgeted_total - offered_total, 2)
    r['per_month_offered']   = round(per_month_offered, 2)
    r['prorata_offered']     = round(prorata_offered, 2)
    r['actual_costing']      = round(actual_costing, 2)

    # Normalise non-serialisable types
    for f in ('id', 'requisition_id', 'created_by'):
        if r.get(f) is not None:
            r[f] = str(r[f])
    for df in ('planned_onboarding_date', 'finance_onboarding_date',
               'salary_budgeted_till', 'offer_date', 'tentative_doj'):
        v = r.get(df)
        if isinstance(v, (date, datetime)):
            r[df] = v.isoformat()
        elif v is not None:
            r[df] = _to_date(v).isoformat() if _to_date(v) else None
    return r


# ── Pipeline sync (imported by pipeline_api) ──────────────────────────────────

def sync_plan_on_advance(
    app_id: str, new_status: str, prev_status: str, req_id: Optional[str]
):
    """
    Called by pipeline_api after every status transition.
    Updates the confirmed-linked hiring plan row for the requisition.
    """
    if not req_id:
        return
    try:
        plan = query_one(
            """SELECT id FROM hiring_plan_rows
               WHERE requisition_id = %s AND link_status = 'confirmed'
               LIMIT 1""",
            [req_id],
        )
        if not plan:
            return
        plan_id = str(plan['id'])

        if new_status == 'offered':
            offer = query_one(
                """SELECT fixed_ctc, variable_ctc
                   FROM offer WHERE application_id = %s
                   ORDER BY created_at DESC LIMIT 1""",
                [app_id],
            )
            cand = query_one(
                """SELECT c.full_name, c.email
                   FROM candidate c
                   JOIN application a ON a.candidate_id = c.id
                   WHERE a.id = %s""",
                [app_id],
            )
            query(
                """UPDATE hiring_plan_rows
                   SET hiring_status   = 'Offered',
                       offered_fixed   = %s,
                       offered_variable = %s,
                       employee_name   = %s,
                       candidate_email = %s,
                       offer_date      = %s,
                       updated_at      = now()
                   WHERE id = %s""",
                [
                    _safe_float((offer or {}).get('fixed_ctc')),
                    _safe_float((offer or {}).get('variable_ctc')),
                    (cand or {}).get('full_name'),
                    (cand or {}).get('email'),
                    date.today(),
                    plan_id,
                ],
                fetch=False,
            )

        elif prev_status == 'offered' and new_status in ('rejected', 'on_hold'):
            # Revert to Open Position only if still showing Offered
            query(
                """UPDATE hiring_plan_rows
                   SET hiring_status   = 'Open Position',
                       offered_fixed   = 0,
                       offered_variable = 0,
                       employee_name   = NULL,
                       candidate_email = NULL,
                       offer_date      = NULL,
                       updated_at      = now()
                   WHERE id = %s AND hiring_status = 'Offered'""",
                [plan_id],
                fetch=False,
            )
    except Exception as exc:
        print(f"[hiring-plan] sync error: {exc}")
        try:
            from ..services.activity_log import log_activity
            log_activity(
                "hiring_plan", "sync_failed",
                entity_id=req_id, requisition_id=req_id, application_id=app_id,
                actor_id=None, actor_role="system",
                detail={"new_status": new_status, "prev_status": prev_status, "error": str(exc)},
            )
        except Exception:
            pass


# ── Header map for xlsx import ────────────────────────────────────────────────

_HEADER_MAP = {
    'Company Entity':                              'company_entity',
    'Quarter':                                     'quarter',
    'Exact Date for Month of Onboarding Required': 'planned_onboarding_date',
    'Requisition ID':                              '__req_code',
    'Role Name':                                   'role_name',
    'BU/ Unit':                                    'bu',
    'Function':                                    'function',
    'Sub BU':                                      'sub_bu',
    'Project Name':                                'project_name',
    'Employment Type':                             'employment_type',
    'Billable/Non-Billable':                       'billable',
    'SOW Received (Yes/No)':                       'sow_received',
    'Resource Required in (OPEX/CAPEX)':           'capex_opex',
    'Capex/Opex on track?':                        'capex_opex_on_track',
    'On-Roll/Off-Roll':                            'on_off_roll',
    'Count':                                       'headcount',
    'Priority':                                    'priority',
    'Band':                                        'band',
    'Experience':                                  'experience',
    'Market Salary Range (Subject to change)':     'market_salary_range',
    'Location':                                    'location',
    'Budgeted Fixed Salary':                       'budgeted_fixed',
    'Budgeted Variable Salary':                    'budgeted_variable',
    'Asset':                                       'asset',
    'Salary Budgeted till':                        'salary_budgeted_till',
    'Hiring Status':                               'hiring_status',
    'Replacement For':                             'replacement_for',
    'AIPLCode':                                    'aipl_code',
    'Employee Name':                               'employee_name',
    'Offered Fixed':                               'offered_fixed',
    'Offered Variable':                            'offered_variable',
    'TA Working on the position':                  'ta_owner',
    'Source of Hire':                              'source_of_hire',
    'Candiate Email':                              'candidate_email',  # intentional sheet typo
    'Offer Date':                                  'offer_date',
    'Tentative Date of Joining':                   'tentative_doj',
    'Remarks':                                     'remarks',
}

# Computed-only columns in the sheet — skip them on import (we recalculate)
_SKIP_HEADERS = {
    'Budgeted Total Salary', 'Months', 'Per Month', 'Pro Rata Budget',
    'Offered Total', 'Budget Offered Diff', 'Per Month Offered',
    'Prorata Offered', 'Actual Costing',
}


# ── Filter helper ─────────────────────────────────────────────────────────────

def _where(fy, quarter, bu, status, priority, tenant_id=None) -> tuple:
    clauses, params = [], []
    if tenant_id:
        clauses.append("h.tenant_id = %s"); params.append(tenant_id)
    if fy:
        clauses.append("h.fiscal_year = %s"); params.append(fy)
    if quarter:
        clauses.append("h.quarter = %s"); params.append(quarter)
    if bu:
        clauses.append("h.bu ILIKE %s"); params.append(f"%{bu}%")
    if status:
        clauses.append("h.hiring_status = %s"); params.append(status)
    if priority:
        clauses.append("h.priority ILIKE %s"); params.append(priority)
    w = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return w, params


# ── Pydantic models ───────────────────────────────────────────────────────────

class HPRowIn(BaseModel):
    fiscal_year:             Optional[str]   = None
    quarter:                 Optional[str]   = None
    company_entity:          Optional[str]   = None
    finance_onboarding_date: Optional[str]   = None
    planned_onboarding_date: Optional[str]   = None
    requisition_id:          Optional[str]   = None
    link_status:             Optional[str]   = 'unlinked'
    role_name:               Optional[str]   = None
    bu:                      Optional[str]   = None
    function:                Optional[str]   = None
    sub_bu:                  Optional[str]   = None
    project_name:            Optional[str]   = None
    employment_type:         Optional[str]   = None
    billable:                Optional[str]   = None
    sow_received:            Optional[str]   = None
    capex_opex:              Optional[str]   = None
    capex_opex_on_track:     Optional[str]   = None
    on_off_roll:             Optional[str]   = None
    headcount:               Optional[int]   = 1
    priority:                Optional[str]   = None
    band:                    Optional[str]   = None
    experience:              Optional[str]   = None
    market_salary_range:     Optional[str]   = None
    location:                Optional[str]   = None
    budgeted_fixed:          Optional[float] = 0
    budgeted_variable:       Optional[float] = 0
    asset:                   Optional[str]   = None
    salary_budgeted_till:    Optional[str]   = None
    hiring_status:           Optional[str]   = 'Open Position'
    replacement_for:         Optional[str]   = None
    aipl_code:               Optional[str]   = None
    employee_name:           Optional[str]   = None
    offered_fixed:           Optional[float] = 0
    offered_variable:        Optional[float] = 0
    ta_owner:                Optional[str]   = None
    source_of_hire:          Optional[str]   = None
    candidate_email:         Optional[str]   = None
    offer_date:              Optional[str]   = None
    tentative_doj:           Optional[str]   = None
    remarks:                 Optional[str]   = None


# ── GET /api/hiring-plan ──────────────────────────────────────────────────────

@router.get("")
def list_rows(
    fy:       Optional[str] = Query(None),
    quarter:  Optional[str] = Query(None),
    bu:       Optional[str] = Query(None),
    status:   Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    _require(user)
    w, params = _where(fy, quarter, bu, status, priority, user.get("tenant_id"))
    rows = query(
        f"""SELECT h.*, r.req_code, r.title AS req_title, r.status AS req_status
            FROM hiring_plan_rows h
            LEFT JOIN requisition r ON r.id = h.requisition_id
            {w}
            ORDER BY h.fiscal_year NULLS LAST, h.quarter NULLS LAST,
                     h.bu NULLS LAST, h.role_name NULLS LAST""",
        params,
    ) or []

    computed = [_compute(r) for r in rows]

    totals = {
        "headcount":      sum(int(r.get('headcount') or 0) for r in computed),
        "budgeted_total": round(sum(r['budgeted_total']  for r in computed), 2),
        "prorata_budget": round(sum(r['prorata_budget']  for r in computed), 2),
        "offered_total":  round(sum(r['offered_total']   for r in computed), 2),
        "actual_costing": round(sum(r['actual_costing']  for r in computed), 2),
        "over_budget":    sum(1 for r in computed if r['budget_offered_diff'] < 0),
    }
    return {"rows": computed, "totals": totals}


# ── POST /api/hiring-plan (manual create) ─────────────────────────────────────

@router.post("")
def create_row(body: HPRowIn, user: dict = Depends(get_current_user)):
    _require(user)
    if body.hiring_status not in HP_STATUSES:
        raise HTTPException(400, f"Invalid hiring_status: {body.hiring_status}")
    if body.headcount is not None and body.headcount < 1:
        raise HTTPException(400, "Headcount must be at least 1.")

    new_id = str(uuid.uuid4())
    req_id = body.requisition_id or None
    link   = body.link_status or 'unlinked'
    if req_id:
        # Validate the requisition exists and belongs to the caller's tenant
        if not query_one("SELECT id FROM requisition WHERE id=%s AND tenant_id=%s", [req_id, user.get("tenant_id")]):
            req_id = None; link = 'unlinked'

    query(
        """INSERT INTO hiring_plan_rows (
               id, fiscal_year, quarter, company_entity,
               finance_onboarding_date, planned_onboarding_date,
               requisition_id, link_status,
               role_name, bu, function, sub_bu, project_name,
               employment_type, billable, sow_received, capex_opex, capex_opex_on_track,
               on_off_roll, headcount, priority, band, experience,
               market_salary_range, location,
               budgeted_fixed, budgeted_variable, asset, salary_budgeted_till,
               hiring_status, replacement_for, aipl_code, employee_name,
               offered_fixed, offered_variable, ta_owner, source_of_hire,
               candidate_email, offer_date, tentative_doj, remarks,
               created_by, tenant_id, created_at, updated_at
           ) VALUES (
               %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
               %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
               %s,%s,%s,now(),now()
           )""",
        [
            new_id,
            body.fiscal_year, body.quarter, body.company_entity,
            _to_date(body.finance_onboarding_date),
            _to_date(body.planned_onboarding_date),
            req_id, link,
            body.role_name, body.bu, body.function, body.sub_bu, body.project_name,
            body.employment_type, body.billable, body.sow_received,
            body.capex_opex, body.capex_opex_on_track, body.on_off_roll,
            body.headcount if body.headcount is not None else 1, body.priority, body.band, body.experience,
            body.market_salary_range, body.location,
            body.budgeted_fixed or 0, body.budgeted_variable or 0,
            body.asset, _to_date(body.salary_budgeted_till),
            body.hiring_status or 'Open Position',
            body.replacement_for, body.aipl_code, body.employee_name,
            body.offered_fixed or 0, body.offered_variable or 0,
            body.ta_owner, body.source_of_hire, body.candidate_email,
            _to_date(body.offer_date), _to_date(body.tentative_doj), body.remarks,
            user["sub"], user.get("tenant_id"),
        ],
        fetch=False,
    )
    row = query_one(
        """SELECT h.*, r.req_code, r.title AS req_title, r.status AS req_status
           FROM hiring_plan_rows h
           LEFT JOIN requisition r ON r.id = h.requisition_id
           WHERE h.id=%s""",
        [new_id],
    )
    return _compute(row)


# ── PUT /api/hiring-plan/{id} ─────────────────────────────────────────────────

@router.put("/{row_id}")
def update_row(row_id: str, body: HPRowIn, user: dict = Depends(get_current_user)):
    _require(user)
    existing = query_one("SELECT id FROM hiring_plan_rows WHERE id=%s AND tenant_id=%s", [row_id, user.get("tenant_id")])
    if not existing:
        raise HTTPException(404, "Row not found")
    if body.hiring_status not in HP_STATUSES:
        raise HTTPException(400, f"Invalid hiring_status: {body.hiring_status}")
    if body.headcount is not None and body.headcount < 1:
        raise HTTPException(400, "Headcount must be at least 1.")

    req_id = body.requisition_id or None
    link   = body.link_status or 'unlinked'
    if req_id:
        if not query_one("SELECT id FROM requisition WHERE id=%s AND tenant_id=%s", [req_id, user.get("tenant_id")]):
            req_id = None; link = 'unlinked'

    query(
        """UPDATE hiring_plan_rows SET
               fiscal_year=%s, quarter=%s, company_entity=%s,
               finance_onboarding_date=%s, planned_onboarding_date=%s,
               requisition_id=%s, link_status=%s,
               role_name=%s, bu=%s, function=%s, sub_bu=%s, project_name=%s,
               employment_type=%s, billable=%s, sow_received=%s,
               capex_opex=%s, capex_opex_on_track=%s, on_off_roll=%s,
               headcount=%s, priority=%s, band=%s, experience=%s,
               market_salary_range=%s, location=%s,
               budgeted_fixed=%s, budgeted_variable=%s, asset=%s,
               salary_budgeted_till=%s, hiring_status=%s,
               replacement_for=%s, aipl_code=%s, employee_name=%s,
               offered_fixed=%s, offered_variable=%s, ta_owner=%s,
               source_of_hire=%s, candidate_email=%s,
               offer_date=%s, tentative_doj=%s, remarks=%s,
               updated_at=now()
           WHERE id=%s""",
        [
            body.fiscal_year, body.quarter, body.company_entity,
            _to_date(body.finance_onboarding_date),
            _to_date(body.planned_onboarding_date),
            req_id, link,
            body.role_name, body.bu, body.function, body.sub_bu, body.project_name,
            body.employment_type, body.billable, body.sow_received,
            body.capex_opex, body.capex_opex_on_track, body.on_off_roll,
            body.headcount if body.headcount is not None else 1, body.priority, body.band, body.experience,
            body.market_salary_range, body.location,
            body.budgeted_fixed or 0, body.budgeted_variable or 0,
            body.asset, _to_date(body.salary_budgeted_till),
            body.hiring_status or 'Open Position',
            body.replacement_for, body.aipl_code, body.employee_name,
            body.offered_fixed or 0, body.offered_variable or 0,
            body.ta_owner, body.source_of_hire, body.candidate_email,
            _to_date(body.offer_date), _to_date(body.tentative_doj), body.remarks,
            row_id,
        ],
        fetch=False,
    )
    row = query_one(
        """SELECT h.*, r.req_code, r.title AS req_title, r.status AS req_status
           FROM hiring_plan_rows h
           LEFT JOIN requisition r ON r.id = h.requisition_id
           WHERE h.id=%s""",
        [row_id],
    )
    return _compute(row)


# ── DELETE /api/hiring-plan/{id} ──────────────────────────────────────────────

@router.delete("/{row_id}")
def delete_row(row_id: str, user: dict = Depends(get_current_user)):
    _require(user)
    existing = query_one("SELECT id FROM hiring_plan_rows WHERE id=%s AND tenant_id=%s", [row_id, user.get("tenant_id")])
    if not existing:
        raise HTTPException(404, "Row not found")
    query("DELETE FROM hiring_plan_rows WHERE id=%s", [row_id], fetch=False)
    return {"ok": True}


# ── POST /api/hiring-plan/import ──────────────────────────────────────────────

@router.post("/import")
async def import_sheet(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    _require(user)
    content = await file.read()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, f"Could not read xlsx: {exc}")

    # Row 2 = headers (1-based in openpyxl → min_row=2)
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not all_rows:
        raise HTTPException(400, "Sheet appears empty")

    header_row = all_rows[0]

    # Build col-index → field mapping
    col_map: dict[int, str] = {}
    for ci, h in enumerate(header_row):
        if h is None:
            continue
        h_str = str(h).strip()
        if h_str in _SKIP_HEADERS:
            continue
        # Flexible match for FY column whose header encodes the year
        if h_str.lower().startswith('considered in fy'):
            col_map[ci] = 'fiscal_year'
            continue
        field = _HEADER_MAP.get(h_str)
        if field:
            col_map[ci] = field

    role_col = next((ci for ci, f in col_map.items() if f == 'role_name'), None)

    imported = skipped = suggested = 0
    uid = user["sub"]

    for data_row in all_rows[1:]:  # data from row 3 onwards
        # Stop at first empty Role Name
        role_val = (data_row[role_col] if role_col is not None and role_col < len(data_row)
                    else None)
        if not role_val or str(role_val).strip() == '':
            break

        data: dict = {}
        for ci, field in col_map.items():
            if ci < len(data_row):
                data[field] = data_row[ci]

        # Parse types
        data['role_name'] = str(data.get('role_name') or '').strip()
        if not data['role_name']:
            skipped += 1
            continue

        for df in ('planned_onboarding_date', 'salary_budgeted_till',
                   'offer_date', 'tentative_doj'):
            data[df] = _to_date(data.get(df))

        for nf in ('budgeted_fixed', 'budgeted_variable', 'offered_fixed', 'offered_variable'):
            data[nf] = _safe_float(data.get(nf))

        data['headcount'] = _safe_int(data.get('headcount'))

        hs = str(data.get('hiring_status') or '').strip()
        if hs not in HP_STATUSES:
            hs = 'Open Position'
        data['hiring_status'] = hs

        # Auto-link
        req_code_raw = str(data.pop('__req_code', '') or '').strip()
        req_id   = None
        link_st  = 'unlinked'

        if req_code_raw:
            req = query_one(
                "SELECT id FROM requisition WHERE req_code=%s AND status='open'",
                [req_code_raw],
            )
            if req:
                req_id  = str(req['id'])
                link_st = 'confirmed'

        if not req_id:
            match = query_one(
                """SELECT r.id FROM requisition r
                   LEFT JOIN business_unit b ON b.id = r.bu_id
                   WHERE r.status = 'open'
                     AND LOWER(r.title) = LOWER(%s)
                   LIMIT 1""",
                [data['role_name']],
            )
            if match:
                req_id  = str(match['id'])
                link_st = 'suggested'
                suggested += 1

        try:
            query(
                """INSERT INTO hiring_plan_rows (
                       id, fiscal_year, quarter, company_entity,
                       planned_onboarding_date,
                       requisition_id, link_status,
                       role_name, bu, function, sub_bu, project_name,
                       employment_type, billable, sow_received,
                       capex_opex, capex_opex_on_track, on_off_roll,
                       headcount, priority, band, experience,
                       market_salary_range, location,
                       budgeted_fixed, budgeted_variable, asset, salary_budgeted_till,
                       hiring_status, replacement_for, aipl_code, employee_name,
                       offered_fixed, offered_variable,
                       ta_owner, source_of_hire, candidate_email,
                       offer_date, tentative_doj, remarks,
                       created_by, created_at, updated_at
                   ) VALUES (
                       gen_random_uuid(),%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                       %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now(),now()
                   )""",
                [
                    data.get('fiscal_year'), data.get('quarter'), data.get('company_entity'),
                    data.get('planned_onboarding_date'),
                    req_id, link_st,
                    data['role_name'], data.get('bu'), data.get('function'),
                    data.get('sub_bu'), data.get('project_name'),
                    data.get('employment_type'), data.get('billable'), data.get('sow_received'),
                    data.get('capex_opex'), data.get('capex_opex_on_track'), data.get('on_off_roll'),
                    data['headcount'], data.get('priority'), data.get('band'),
                    data.get('experience'), data.get('market_salary_range'), data.get('location'),
                    data['budgeted_fixed'], data['budgeted_variable'],
                    data.get('asset'), data.get('salary_budgeted_till'),
                    data['hiring_status'], data.get('replacement_for'), data.get('aipl_code'),
                    data.get('employee_name'),
                    data['offered_fixed'], data['offered_variable'],
                    data.get('ta_owner'), data.get('source_of_hire'), data.get('candidate_email'),
                    data.get('offer_date'), data.get('tentative_doj'), data.get('remarks'),
                    uid,
                ],
                fetch=False,
            )
            imported += 1
        except Exception as exc:
            print(f"[hiring-plan import] row skipped: {exc}")
            skipped += 1

    return {"imported": imported, "skipped": skipped, "suggested_links": suggested}


# ── POST /api/hiring-plan/{id}/create-requisition ────────────────────────────

@router.post("/{row_id}/create-requisition")
def create_req_from_plan(row_id: str, user: dict = Depends(get_current_user)):
    _require(user)
    tenant_id = user.get("tenant_id")
    hp = query_one("SELECT * FROM hiring_plan_rows WHERE id=%s AND tenant_id=%s", [row_id, tenant_id])
    if not hp:
        raise HTTPException(404, "Plan row not found")

    # Try to resolve bu_id from bu name, scoped to the caller's own tenant
    bu_id = None
    if hp.get('bu'):
        bu_row = query_one(
            """SELECT bu.id FROM business_unit bu
               JOIN group_company gc ON gc.id = bu.company_id
               WHERE gc.tenant_id = %s AND LOWER(bu.name) ILIKE LOWER(%s) LIMIT 1""",
            [tenant_id, hp['bu']],
        )
        if bu_row:
            bu_id = str(bu_row['id'])

    # Parse min/max experience from free text (e.g. "3-5 years", "5+")
    min_exp = max_exp = None
    exp_text = str(hp.get('experience') or '')
    m = re.search(r'(\d+)\s*[-–]\s*(\d+)', exp_text)
    if m:
        min_exp, max_exp = float(m.group(1)), float(m.group(2))
    else:
        m2 = re.search(r'(\d+)', exp_text)
        if m2:
            min_exp = float(m2.group(1))

    # Map priority text to DB enum
    prio_map = {
        'critical': 'critical', 'high': 'high', 'medium': 'medium', 'low': 'low',
        'p1': 'critical', 'p2': 'high', 'p3': 'medium',
    }
    priority = prio_map.get(str(hp.get('priority') or '').lower())

    # Auto-generate req_code, retrying on a concurrent-request collision instead
    # of letting the UNIQUE constraint violation crash the request with a raw
    # 500 (same fix as pipeline_api.py's create_requisition).
    import psycopg2
    req_id = str(uuid.uuid4())
    inserted = False
    for _attempt in range(5):
        cnt_row = query_one("SELECT COUNT(*) AS n FROM requisition")
        next_n  = int((cnt_row or {}).get('n') or 0) + 1
        req_code = f"REQ-{next_n:04d}"
        while query_one("SELECT 1 FROM requisition WHERE req_code=%s", [req_code]):
            next_n += 1
            req_code = f"REQ-{next_n:04d}"
        try:
            query(
                """INSERT INTO requisition (
                       id, req_code, title, bu_id, roll_type, openings,
                       budgeted_fixed, budgeted_variable,
                       min_experience, max_experience,
                       grade_level, priority, project, status,
                       created_by, tenant_id, created_at, opened_at
                   ) VALUES (
                       %s,%s,%s,%s,'on_roll',%s,%s,%s,%s,%s,%s,%s,%s,'open',%s,%s,now(),now()
                   )""",
                [
                    req_id, req_code, hp.get('role_name') or 'Untitled',
                    bu_id,
                    hp.get('headcount') or 1,
                    _safe_float(hp.get('budgeted_fixed')),
                    _safe_float(hp.get('budgeted_variable')),
                    min_exp, max_exp,
                    hp.get('band'), priority, hp.get('project_name'),
                    user["sub"], tenant_id,
                ],
                fetch=False,
            )
            inserted = True
            break
        except psycopg2.errors.UniqueViolation:
            continue
    if not inserted:
        raise HTTPException(409, "Could not generate a unique requisition code — please retry.")

    # Auto-assign the creating recruiter as owner -- a ta_manager can also be
    # an individual contributor (personally own/work a requisition), not just
    # a recruiter, so their own analytics/reports aren't left empty (same fix
    # as pipeline_api.py's create_requisition).
    if user["role"] in ("recruiter", "ta_manager"):
        query(
            """INSERT INTO requisition_recruiter
               (requisition_id, recruiter_id, is_owner, assigned_by)
               VALUES (%s,%s,true,%s) ON CONFLICT DO NOTHING""",
            [req_id, user["sub"], user["sub"]], fetch=False,
        )

    # Link plan row
    query(
        "UPDATE hiring_plan_rows SET requisition_id=%s, link_status='confirmed', updated_at=now() WHERE id=%s",
        [req_id, row_id], fetch=False,
    )

    return {"ok": True, "requisition_id": req_id, "req_code": req_code}


# ── GET /api/hiring-plan/summary ─────────────────────────────────────────────

@router.get("/summary")
def hiring_plan_summary(
    fy: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    _require(user)
    conds = ["tenant_id = %s"]
    params: list = [user.get("tenant_id")]
    if fy:
        conds.append("fiscal_year = %s")
        params.append(fy)
    fy_where = "WHERE " + " AND ".join(conds)

    # By BU
    by_bu = query(
        f"""SELECT bu,
                   SUM(headcount)                                                AS demand,
                   SUM(headcount) FILTER (WHERE hiring_status IN ('Offered','Joined')) AS fulfilled,
                   SUM(headcount) FILTER (WHERE hiring_status = 'Hold')          AS on_hold,
                   SUM(headcount) FILTER (WHERE hiring_status = 'Open Position') AS open
            FROM hiring_plan_rows {fy_where}
            GROUP BY bu ORDER BY demand DESC NULLS LAST""",
        params,
    ) or []

    # By quarter
    by_quarter = query(
        f"""SELECT quarter,
                   SUM(headcount)                                                AS demand,
                   SUM(headcount) FILTER (WHERE hiring_status IN ('Offered','Joined')) AS fulfilled,
                   SUM(headcount) FILTER (WHERE hiring_status = 'Hold')          AS on_hold,
                   SUM(headcount) FILTER (WHERE hiring_status = 'Open Position') AS open
            FROM hiring_plan_rows {fy_where}
            GROUP BY quarter ORDER BY quarter NULLS LAST""",
        params,
    ) or []

    # By priority
    by_priority = query(
        f"""SELECT priority,
                   SUM(headcount) AS demand,
                   SUM(headcount) FILTER (WHERE hiring_status IN ('Offered','Joined')) AS fulfilled
            FROM hiring_plan_rows {fy_where}
            GROUP BY priority ORDER BY demand DESC NULLS LAST""",
        params,
    ) or []

    # Budget KPIs — computed in Python (prorata requires date math)
    all_rows = query(
        f"SELECT * FROM hiring_plan_rows {fy_where}", params
    ) or []
    computed = [_compute(r) for r in all_rows]

    kpis = {
        "planned_headcount":  sum(int(r.get('headcount') or 0) for r in computed),
        "fulfilled":          sum(1 for r in computed if r.get('hiring_status') in ('Offered','Joined')),
        "open":               sum(1 for r in computed if r.get('hiring_status') == 'Open Position'),
        "on_hold":            sum(1 for r in computed if r.get('hiring_status') == 'Hold'),
        "total_budgeted":     round(sum(r['budgeted_total']  for r in computed), 2),
        "total_prorata":      round(sum(r['prorata_budget']  for r in computed), 2),
        "total_offered":      round(sum(r['offered_total']   for r in computed), 2),
        "total_actual":       round(sum(r['actual_costing']  for r in computed), 2),
        "over_budget_count":  sum(1 for r in computed if r['budget_offered_diff'] < 0),
    }

    return {
        "kpis":        kpis,
        "by_bu":       by_bu,
        "by_quarter":  by_quarter,
        "by_priority": by_priority,
    }


# ── GET /api/hiring-plan/excel ────────────────────────────────────────────────

_EXCEL_COLS = [
    ('Company Entity',              'company_entity'),
    ('Fiscal Year',                 'fiscal_year'),
    ('Quarter',                     'quarter'),
    ('Planned Onboarding Date',     'planned_onboarding_date'),
    ('Finance Onboarding Date',     'finance_onboarding_date'),
    ('Requisition ID',              'req_code'),
    ('Link Status',                 'link_status'),
    ('Role Name',                   'role_name'),
    ('BU',                          'bu'),
    ('Function',                    'function'),
    ('Sub BU',                      'sub_bu'),
    ('Project Name',                'project_name'),
    ('Employment Type',             'employment_type'),
    ('Billable',                    'billable'),
    ('SOW Received',                'sow_received'),
    ('CAPEX/OPEX',                  'capex_opex'),
    ('CAPEX/OPEX On Track',         'capex_opex_on_track'),
    ('On/Off Roll',                 'on_off_roll'),
    ('Headcount',                   'headcount'),
    ('Priority',                    'priority'),
    ('Band',                        'band'),
    ('Experience',                  'experience'),
    ('Market Salary Range',         'market_salary_range'),
    ('Location',                    'location'),
    ('Budgeted Fixed',              'budgeted_fixed'),
    ('Budgeted Variable',           'budgeted_variable'),
    ('Budgeted Total',              'budgeted_total'),
    ('Asset',                       'asset'),
    ('Salary Budgeted Till',        'salary_budgeted_till'),
    ('Months',                      'months'),
    ('Per Month Budget',            'per_month_budget'),
    ('Pro-Rata Budget',             'prorata_budget'),
    ('Hiring Status',               'hiring_status'),
    ('Replacement For',             'replacement_for'),
    ('AIPL Code',                   'aipl_code'),
    ('Employee Name',               'employee_name'),
    ('Offered Fixed',               'offered_fixed'),
    ('Offered Variable',            'offered_variable'),
    ('Offered Total',               'offered_total'),
    ('Budget vs Offered Diff',      'budget_offered_diff'),
    ('Per Month Offered',           'per_month_offered'),
    ('Pro-Rata Offered',            'prorata_offered'),
    ('Actual Costing',              'actual_costing'),
    ('TA Owner',                    'ta_owner'),
    ('Source of Hire',              'source_of_hire'),
    ('Candidate Email',             'candidate_email'),
    ('Offer Date',                  'offer_date'),
    ('Tentative DOJ',               'tentative_doj'),
    ('Remarks',                     'remarks'),
]


@router.get("/excel")
def export_excel(
    fy:       Optional[str] = Query(None),
    quarter:  Optional[str] = Query(None),
    bu:       Optional[str] = Query(None),
    status:   Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    _require(user)
    import openpyxl

    w, params = _where(fy, quarter, bu, status, priority, user.get("tenant_id"))
    rows = query(
        f"""SELECT h.*, r.req_code, r.title AS req_title
            FROM hiring_plan_rows h
            LEFT JOIN requisition r ON r.id = h.requisition_id
            {w} ORDER BY h.fiscal_year, h.quarter, h.bu""",
        params,
    ) or []
    computed = [_compute(r) for r in rows]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    excel_export.sheet_from_rows(wb, "Hiring Plan", computed, columns=_EXCEL_COLS)
    excel_export.build_summary_sheet(
        wb,
        title=f"Hiring Plan{' — ' + fy if fy else ''}",
        generated_by=user.get("name") or user.get("email") or "",
        generated_at=datetime.now(),
        filters_applied=[
            f for f in [
                {"key": "fy", "op": "=", "value": fy} if fy else None,
                {"key": "quarter", "op": "=", "value": quarter} if quarter else None,
                {"key": "bu", "op": "=", "value": bu} if bu else None,
                {"key": "status", "op": "=", "value": status} if status else None,
                {"key": "priority", "op": "=", "value": priority} if priority else None,
            ] if f
        ],
        rows=computed,
    )
    fname = f"enternly_hiring_plan{'_'+fy if fy else ''}.xlsx"
    return excel_export.stream_workbook(wb, fname)
