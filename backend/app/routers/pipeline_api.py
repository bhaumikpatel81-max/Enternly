"""
New pipeline API endpoints: dashboard, requisitions CRUD,
kanban, candidates, interviews, hiring-manager review.
"""
import json as _json
import os as _os
import re as _re
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, File, Form, Query, Response, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

from ..db import query, query_one, transaction, tx_exec
from ..auth_utils import get_current_user
from ..services import connectors
from ..services.gamification import award as _gam_award
from ..services.activity_log import log_activity

router = APIRouter(prefix="/api", tags=["pipeline"])

from ..services.sla import (
    PIPELINE_STAGES, PIPELINE_STAGE_LABELS, NEXT_STAGE, TERMINAL,
    STAGE_SLA_KEY, load_config, compute_rag, derive_pending_from,
)
from .hiring_plan_api import sync_plan_on_advance
from .scheduling_api import (
    create_schedule_request_tx as _create_schedule_request_tx,
    _after_insert_side_effects as _schedule_side_effects,
)
from .nexai_api import _recruiter_owns_req, _application_req_id
from .hrbp_api import scope_requisitions_for_hrbp


def _send_application_rejected_email(app_id: str, user: Optional[dict] = None) -> None:
    """Best-effort candidate-facing rejection email -- never raises, so a
    mail hiccup can never block the status change/offer-cancel action that
    triggered it. Shared by every path that moves an application to
    'rejected' (generic reject, screening-decision reject, offer cancel).
    Uses the admin-editable "application_rejected" template (Settings > Email
    Templates) wrapped in the same branded HTML layout as every other
    candidate email (confirmation, interview, etc.).

    Every outcome (sent / skipped / failed) is written to activity_log --
    this used to be a bare print() that only existed in stdout, so there was
    no way after the fact to tell whether a "rejected" candidate actually got
    an email or the send silently no-op'd (e.g. connectors.send_email()
    returns {"sent": False} rather than raising when SMTP isn't configured,
    which a caller that ignores the return value would never notice)."""
    try:
        ctx = query_one(
            """SELECT c.full_name AS candidate_name, c.email AS candidate_email,
                      r.id AS requisition_id, r.title AS job_title,
                      r.tenant_id AS tenant_id,
                      gc.name AS company
               FROM application a
               JOIN candidate c ON c.id = a.candidate_id
               JOIN requisition r ON r.id = a.requisition_id
               JOIN business_unit bu ON bu.id = r.bu_id
               JOIN group_company gc ON gc.id = bu.company_id
               WHERE a.id = %s""",
            [app_id],
        )
        if not ctx or not ctx.get("candidate_email"):
            log_activity(
                "application", "rejection_email_skipped",
                entity_id=app_id, application_id=app_id,
                requisition_id=ctx and str(ctx.get("requisition_id")),
                actor_id=user and user.get("sub"), actor_role=user and user.get("role"),
                detail={"reason": "no candidate email on file"},
            )
            return
        import html as _html
        from ..services.email_templates import render_template
        from ..services.email_layout import build_branded_email

        candidate_name = ctx["candidate_name"] or "there"
        job_title = ctx["job_title"] or "the role"
        subject, body = render_template(
            "application_rejected",
            {"candidate_name": candidate_name, "job_title": job_title},
            req_id=str(ctx["requisition_id"]), actor=user,
        )
        html_body = build_branded_email(
            eyebrow="Application Tracking System",
            hero_title_html="Application<br>Update.",
            hero_subtitle=f"Hi {_html.escape(candidate_name)}, thank you for your interest in the {_html.escape(job_title)} role.",
            hero_footer_label=job_title, hero_footer_value=ctx.get("company"),
            detail_cells=[("Candidate", candidate_name), ("Position", job_title)],
            about_text=body,
            about_heading=None,
            footer_note="We wish you the best in your career search.",
        )
        result = connectors.send_email(ctx["candidate_email"], subject, body, html=html_body, tenant_id=ctx.get("tenant_id"))
        if not result.get("sent"):
            # send_email() returns {"sent": False} rather than raising when
            # SMTP isn't configured -- treat that the same as a real failure.
            raise RuntimeError("email not sent — SMTP not configured (stub mode)")
        log_activity(
            "application", "rejection_email_sent",
            entity_id=app_id, application_id=app_id,
            requisition_id=str(ctx["requisition_id"]),
            actor_id=user and user.get("sub"), actor_role=user and user.get("role"),
            detail={"to": ctx["candidate_email"]},
        )
    except Exception as exc:
        print(f"[pipeline] rejection email failed for application {app_id}: {exc}")
        try:
            log_activity(
                "application", "rejection_email_failed",
                entity_id=app_id, application_id=app_id,
                actor_id=user and user.get("sub"), actor_role=user and user.get("role"),
                detail={"error": str(exc)[:500]},
            )
        except Exception:
            pass


# ─── helpers ──────────────────────────────────────────────────────────────────

def _is_recruiter_scoped(role: str) -> bool:
    return role == "recruiter"


def _schedule_request_for_round_tx(cur, req_id: str, app_id: str, round_seq: int, actor_id: str) -> tuple:
    """Every round a candidate enters (any round_type, auto or manual) must get
    an interview_schedule_request — inserted on the SAME transaction cursor as
    the stage move, so a failure here rolls back the stage change too instead
    of silently leaving a candidate parked in an interview stage with nothing
    scheduled. is_auto/round_type no longer gate whether a request is created;
    they're just round metadata now."""
    rc = query_one(
        "SELECT id FROM round_config WHERE requisition_id=%s AND sequence=%s",
        [req_id, round_seq],
    )
    if not rc:
        raise HTTPException(400, f"No round configured at sequence {round_seq} for this requisition")
    return _create_schedule_request_tx(cur, app_id, round_config_id=str(rc["id"]), created_by=actor_id)


def _deny_hrbp(user: dict) -> None:
    """HRBP's only reachable read surface is /api/hrbp/* (own requisitions +
    per-requisition candidate status). Every broader endpoint here would leak
    scores/PII beyond that, so it's explicitly rejected rather than merely
    left off an allowlist."""
    if user.get("role") == "hrbp":
        raise HTTPException(403, "Not available to the HRBP role -- use My Requisitions")


# ─── Cross-role pipeline status (Phase 4) ─────────────────────────────────────

@router.get("/pipeline/status")
def pipeline_status(user: dict = Depends(get_current_user)):
    """Scoped, read-only view of where every candidate is and who's blocking
    progress -- recruiter/HM/HRBP/TA manager each get exactly their own slice,
    resolved server-side from existing relationships. Never trust a
    client-supplied scope; there isn't one to trust here on purpose."""
    role = user["role"]
    where, params = "", []
    if role == "recruiter":
        where = "AND r.id IN (SELECT requisition_id FROM requisition_recruiter WHERE recruiter_id = %s)"
        params = [user["sub"]]
    elif role == "hiring_manager":
        where = "AND r.hiring_manager_id = %s"
        params = [user["sub"]]
    elif role == "hrbp":
        hrbp_where, hrbp_params = scope_requisitions_for_hrbp(user)
        where, params = f"AND {hrbp_where}", hrbp_params
    elif role == "ta_manager":
        pass  # unscoped -- full pipeline visibility
    else:
        raise HTTPException(403, "Not authorised")

    rows = query(
        f"""SELECT a.id AS application_id, a.status AS app_status,
                   c.full_name AS candidate_name,
                   r.id AS req_id, r.title AS req_title,
                   isr.status AS isr_status
            FROM application a
            JOIN candidate   c ON c.id = a.candidate_id
            JOIN requisition r ON r.id = a.requisition_id
            LEFT JOIN LATERAL (
                SELECT status FROM interview_schedule_request
                WHERE application_id = a.id
                ORDER BY created_at DESC LIMIT 1
            ) isr ON true
            WHERE 1=1 {where}
            ORDER BY r.created_at DESC
            LIMIT 500""",
        params,
    )
    return [
        {
            "application_id": str(row["application_id"]),
            "candidate_name": row["candidate_name"],
            "req_id":         str(row["req_id"]),
            "req_title":      row["req_title"],
            "status":         row["app_status"],
            "status_label":   PIPELINE_STAGE_LABELS.get(row["app_status"], row["app_status"]),
            "pending_from":   derive_pending_from(row["app_status"], row["isr_status"]),
        }
        for row in (rows or [])
    ]


# ─── Dashboard ────────────────────────────────────────────────────────────────

@router.get("/dashboard")
def dashboard(user: dict = Depends(get_current_user)):
    _deny_hrbp(user)
    role = user["role"]
    uid  = user["sub"]

    if role == "recruiter":
        # Only count applications under the recruiter's own requisitions
        app_filter = """
            a.requisition_id IN (
                SELECT requisition_id FROM requisition_recruiter WHERE recruiter_id = %s
            )
        """
        req_filter = """
            r.id IN (
                SELECT requisition_id FROM requisition_recruiter WHERE recruiter_id = %s
            )
        """
        p = [uid]
    else:
        app_filter = "1=1"
        req_filter  = "1=1"
        p = []

    def cnt(extra_where):
        row = query_one(
            f"SELECT COUNT(*) AS n FROM application a WHERE {app_filter} AND ({extra_where})",
            p,
        )
        return int(row["n"]) if row else 0

    open_reqs = query_one(
        f"SELECT COUNT(*) AS n FROM requisition r WHERE {req_filter} AND r.status='open' AND COALESCE(r.approval_status,'approved')='approved'",
        p,
    )

    counts = {
        "open_reqs":         int(open_reqs["n"]) if open_reqs else 0,
        "apps_received":     cnt("1=1"),
        "under_screening":   cnt("a.status='screen'"),
        "screening_cleared": cnt("a.status IN ('shortlisted','nexai_bot')"),
        "ai_interview":      cnt("a.bot_score IS NOT NULL"),
        "panel_interview":   cnt("a.status='interview'"),
        "selected":          cnt("a.status='documentation'"),
        "offer_stage":       cnt("a.status='offered'"),
        "joined":            cnt("a.status='hired'"),
    }

    # Average days to hire (stage_event: applied → joined)
    ath = query_one(
        """
        SELECT ROUND(
            AVG(EXTRACT(EPOCH FROM (e2.occurred_at - e1.occurred_at)) / 86400)::numeric, 1
        ) AS avg_days
        FROM stage_event e1
        JOIN stage_event e2 ON e2.application_id = e1.application_id
        WHERE e1.to_status = 'applied' AND e2.to_status = 'hired'
        """,
        [],
    )
    counts["avg_days_to_hire"] = float(ath["avg_days"]) if ath and ath["avg_days"] else None

    # Gender split (global or scoped)
    if role == "recruiter":
        gender = query(
            """
            SELECT c.gender, COUNT(*) AS n
            FROM application a
            JOIN candidate c ON c.id = a.candidate_id
            WHERE a.requisition_id IN (
                SELECT requisition_id FROM requisition_recruiter WHERE recruiter_id = %s
            )
            GROUP BY c.gender
            """,
            [uid],
        )
    else:
        gender = query(
            "SELECT gender, COUNT(*) AS n FROM candidate GROUP BY gender",
            [],
        )
    counts["gender_split"] = gender

    # Recent requisitions (scoped)
    if role == "recruiter":
        reqs = query(
            """
            SELECT r.id, r.title, r.status, b.code AS band, bu.name AS business_unit,
                   (SELECT COUNT(*) FROM application WHERE requisition_id = r.id) AS in_pipeline,
                   (SELECT COUNT(*) FROM round_config  WHERE requisition_id = r.id) AS levels
            FROM requisition r
            JOIN requisition_recruiter rr ON rr.requisition_id = r.id AND rr.recruiter_id = %s
            JOIN band b          ON b.id = r.band_id
            JOIN business_unit bu ON bu.id = r.bu_id
            ORDER BY r.created_at DESC LIMIT 10
            """,
            [uid],
        )
    elif role == "hiring_manager":
        reqs = query(
            """
            SELECT r.id, r.title, r.status, b.code AS band, bu.name AS business_unit,
                   (SELECT COUNT(*) FROM application WHERE requisition_id = r.id) AS in_pipeline,
                   (SELECT COUNT(*) FROM round_config  WHERE requisition_id = r.id) AS levels
            FROM requisition r
            JOIN band b          ON b.id = r.band_id
            JOIN business_unit bu ON bu.id = r.bu_id
            WHERE r.hiring_manager_id = %s
            ORDER BY r.created_at DESC LIMIT 10
            """,
            [uid],
        )
    else:
        reqs = query(
            """
            SELECT r.id, r.title, r.status, b.code AS band, bu.name AS business_unit,
                   (SELECT COUNT(*) FROM application WHERE requisition_id = r.id) AS in_pipeline,
                   (SELECT COUNT(*) FROM round_config  WHERE requisition_id = r.id) AS levels
            FROM requisition r
            JOIN band b          ON b.id = r.band_id
            JOIN business_unit bu ON bu.id = r.bu_id
            ORDER BY r.created_at DESC LIMIT 10
            """,
            [],
        )
    counts["recent_reqs"] = reqs

    # ─── NexAI data ────────────────────────────────────────────────────────────

    _NX_SUMMARY_COLS = """
        COUNT(*)                                                           AS total,
        COUNT(*) FILTER (WHERE ns.status = 'completed')                    AS completed,
        COUNT(*) FILTER (WHERE ns.status = 'failed')                       AS failed,
        COUNT(*) FILTER (WHERE ns.status IN ('pending','in_progress'))     AS pending,
        ROUND(AVG(ns.raw_score) FILTER (WHERE ns.status='completed')
              ::numeric, 1)                                                AS avg_score,
        COUNT(*) FILTER (WHERE ns.raw_score >= 70 AND ns.status='completed') AS high_scorers,
        COUNT(*) FILTER (WHERE ns.raw_score <  40 AND ns.status='completed') AS low_scorers,
        ROUND(
          COALESCE(
            COUNT(*) FILTER (WHERE ns.raw_score >= 50 AND ns.status='completed')
            ::numeric /
            NULLIF(COUNT(*) FILTER (WHERE ns.status='completed'), 0),
          0) * 100, 1
        )                                                                  AS pass_rate
    """

    if role == "recruiter":
        nx_where = """
            JOIN requisition r2 ON r2.id = ns.requisition_id
            JOIN requisition_recruiter rr2
                 ON rr2.requisition_id = r2.id AND rr2.recruiter_id = %s
        """
        nx_params = [uid]
        nx_dist_where = f"""
            WHERE ns.status = 'completed' AND ns.raw_score IS NOT NULL
              AND ns.requisition_id IN (
                  SELECT requisition_id FROM requisition_recruiter WHERE recruiter_id = %s
              )
        """
        nx_dist_params = [uid]
        nx_recent_where = f"""
            JOIN requisition ri ON ri.id = ns.requisition_id
            JOIN requisition_recruiter rir
                 ON rir.requisition_id = ri.id AND rir.recruiter_id = %s
        """
        nx_recent_params = [uid]
    else:
        nx_where = ""
        nx_params = []
        nx_dist_where = "WHERE ns.status = 'completed' AND ns.raw_score IS NOT NULL"
        nx_dist_params = []
        nx_recent_where = ""
        nx_recent_params = []

    if role in ("recruiter", "ta_manager"):
        nx_row = query_one(
            f"SELECT {_NX_SUMMARY_COLS} FROM nexai_session ns {nx_where}",
            nx_params,
        )
        counts["nexai_summary"] = dict(nx_row) if nx_row else {}

        counts["nexai_score_dist"] = query(
            f"""
            SELECT
              CASE
                WHEN raw_score >= 80 THEN '80-100'
                WHEN raw_score >= 60 THEN '60-79'
                WHEN raw_score >= 40 THEN '40-59'
                WHEN raw_score >= 20 THEN '20-39'
                ELSE '0-19'
              END AS bucket,
              CASE
                WHEN raw_score >= 80 THEN 5
                WHEN raw_score >= 60 THEN 4
                WHEN raw_score >= 40 THEN 3
                WHEN raw_score >= 20 THEN 2
                ELSE 1
              END AS sort_ord,
              COUNT(*) AS n
            FROM nexai_session ns
            {nx_dist_where}
            GROUP BY bucket, sort_ord
            ORDER BY sort_ord
            """,
            nx_dist_params,
        )

        counts["nexai_recent"] = query(
            f"""
            SELECT ns.id, ns.raw_score, ns.status,
                   ns.created_at, ns.completed_at,
                   c.full_name AS candidate_name,
                   r.title     AS req_title
            FROM nexai_session ns
            JOIN application a ON a.id = ns.application_id
            JOIN candidate   c ON c.id = a.candidate_id
            JOIN requisition r ON r.id = ns.requisition_id
            {nx_recent_where}
            ORDER BY ns.created_at DESC LIMIT 10
            """,
            nx_recent_params,
        )

    if role == "ta_manager":
        counts["nexai_by_recruiter"] = query(
            """
            SELECT u.full_name AS recruiter_name,
                   COUNT(ns.id)                                                   AS total,
                   COUNT(ns.id) FILTER (WHERE ns.status='completed')              AS completed,
                   ROUND(AVG(ns.raw_score) FILTER
                         (WHERE ns.status='completed')::numeric, 1)               AS avg_score,
                   COUNT(ns.id) FILTER
                         (WHERE ns.raw_score >= 70 AND ns.status='completed')     AS high_scorers
            FROM app_user u
            LEFT JOIN requisition_recruiter rr ON rr.recruiter_id = u.id
            LEFT JOIN nexai_session ns ON ns.requisition_id = rr.requisition_id
            WHERE u.role IN ('recruiter','ta_manager') AND u.is_active = true
            GROUP BY u.id, u.full_name
            ORDER BY avg_score DESC NULLS LAST, u.full_name
            """,
            [],
        )

    # Recruiter load panel (ta_manager / admin only)
    if role in ("ta_manager", "admin"):
        counts["recruiter_load"] = query("SELECT * FROM v_recruiter_load", [])

    # TA Manager: hiring manager overview
    if role == "ta_manager":
        counts["hiring_manager_stats"] = query(
            """
            SELECT u.id AS hm_id, u.full_name, u.email,
                   COUNT(DISTINCT r.id) AS assigned_reqs,
                   SUM(CASE WHEN a.status = 'interview'
                                 AND (a.hm_feedback IS NULL OR a.hm_feedback = '')
                            THEN 1 ELSE 0 END)              AS pending_reviews,
                   COUNT(DISTINCT CASE WHEN a.hm_feedback IS NOT NULL
                                            AND a.hm_feedback != ''
                                       THEN a.id END)       AS reviews_done
            FROM app_user u
            LEFT JOIN requisition r  ON r.hiring_manager_id = u.id
            LEFT JOIN application a  ON a.requisition_id    = r.id
            WHERE u.role = 'hiring_manager' AND u.is_active = true
            GROUP BY u.id, u.full_name, u.email
            ORDER BY pending_reviews DESC, u.full_name
            """,
            [],
        )

    # Hiring manager: profiles + interviews + nexai + skills + time data
    if role == "hiring_manager":
        counts["profiles_to_review"] = query(
            """
            SELECT a.id, c.full_name, r.title AS req_title,
                   a.combined_score, a.match_score, a.status
            FROM application a
            JOIN candidate  c ON c.id  = a.candidate_id
            JOIN requisition r ON r.id = a.requisition_id
            WHERE r.hiring_manager_id = %s
              AND a.status = 'interview'
              AND (a.hm_feedback IS NULL OR a.hm_feedback = '')
            ORDER BY a.combined_score DESC NULLS LAST
            LIMIT 20
            """,
            [uid],
        )
        counts["my_interviews"] = query(
            """
            SELECT i.id, i.scheduled_at, i.mode, i.duration_min,
                   COALESCE(i.status, 'scheduled') AS status,
                   c.full_name  AS candidate_name,
                   r.title      AS req_title,
                   rc.name      AS round_name
            FROM interview i
            JOIN application  a  ON a.id  = i.application_id
            JOIN candidate    c  ON c.id  = a.candidate_id
            JOIN requisition  r  ON r.id  = a.requisition_id
            LEFT JOIN round_config rc ON rc.id = i.round_config_id
            WHERE r.hiring_manager_id = %s
            ORDER BY i.scheduled_at DESC LIMIT 10
            """,
            [uid],
        )
        counts["feedback_outcomes"] = query(
            """
            SELECT COALESCE(NULLIF(a.hm_feedback,''), 'pending') AS outcome,
                   COUNT(*) AS n
            FROM application a
            JOIN requisition r ON r.id = a.requisition_id
            WHERE r.hiring_manager_id = %s
              AND a.status IN ('interview','documentation','offered','hired','rejected')
            GROUP BY COALESCE(NULLIF(a.hm_feedback,''), 'pending')
            ORDER BY n DESC
            """,
            [uid],
        )

        # Interviews conducted + time stats
        itime = query_one(
            """
            SELECT COUNT(DISTINCT i.id)                          AS n,
                   ROUND(AVG(i.duration_min)::numeric, 0)        AS avg_min,
                   ROUND(SUM(i.duration_min)::numeric / 60.0, 1) AS total_hrs
            FROM interview i
            JOIN application a  ON a.id = i.application_id
            JOIN requisition r  ON r.id = a.requisition_id
            WHERE r.hiring_manager_id = %s
            """,
            [uid],
        )
        counts["interviews_conducted"] = int(itime["n"]) if itime else 0
        counts["avg_interview_min"]    = float(itime["avg_min"])   if itime and itime["avg_min"]   else None
        counts["total_interview_hrs"]  = float(itime["total_hrs"]) if itime and itime["total_hrs"] else 0

        # NexAI screening summary for HM's requisitions
        nexai = query_one(
            """
            SELECT
              COUNT(*)                                                  AS total,
              COUNT(*) FILTER (WHERE ns.status = 'completed')          AS completed,
              COUNT(*) FILTER (WHERE ns.status = 'failed')             AS failed,
              COUNT(*) FILTER (WHERE ns.status IN ('pending','in_progress')) AS pending,
              ROUND(AVG(ns.raw_score) FILTER
                    (WHERE ns.status='completed')::numeric, 1)          AS avg_score,
              ROUND(AVG(
                EXTRACT(EPOCH FROM (ns.completed_at - ns.started_at))/60.0
              ) FILTER (WHERE ns.status='completed')::numeric, 1)       AS avg_session_min
            FROM nexai_session ns
            JOIN requisition r ON r.id = ns.requisition_id
            WHERE r.hiring_manager_id = %s
            """,
            [uid],
        )
        counts["nexai_summary"] = dict(nexai) if nexai else {
            "total": 0, "completed": 0, "failed": 0, "pending": 0,
            "avg_score": None, "avg_session_min": None,
        }

        # Skills breakdown: aggregate key_skills from HM's requisitions
        counts["skills_summary"] = query(
            """
            SELECT UNNEST(key_skills) AS skill, COUNT(*) AS n
            FROM requisition
            WHERE hiring_manager_id = %s
              AND key_skills IS NOT NULL AND array_length(key_skills, 1) > 0
            GROUP BY skill
            ORDER BY n DESC, skill
            LIMIT 15
            """,
            [uid],
        )

    return counts


@router.get("/dashboard/excel")
def dashboard_excel(user: dict = Depends(get_current_user)):
    _deny_hrbp(user)
    from datetime import datetime
    import openpyxl
    from ..services import excel_export

    data = dashboard(user=user)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    scalar_rows = []
    list_sheets = []
    for key, val in data.items():
        if isinstance(val, list) and val and isinstance(val[0], dict):
            list_sheets.append((key, val))
        elif isinstance(val, dict):
            for subk, subv in val.items():
                if not isinstance(subv, (list, dict)):
                    scalar_rows.append({"Metric": f"{key}.{subk}", "Value": subv})
        elif not isinstance(val, list):
            scalar_rows.append({"Metric": key, "Value": val})

    if scalar_rows:
        excel_export.sheet_from_rows(wb, "KPI Summary", scalar_rows)
    for key, rows in list_sheets:
        excel_export.sheet_from_rows(wb, key.replace("_", " ").title()[:31], rows)
    if not wb.sheetnames:
        excel_export.sheet_from_rows(wb, "Data", [])

    excel_export.build_summary_sheet(
        wb,
        title=f"Dashboard — {user['role'].replace('_', ' ').title()}",
        generated_by=user.get("name") or user.get("email") or "",
        generated_at=datetime.now(),
        filters_applied=[],
        rows=scalar_rows,
        measures_meta=[{"key": "Value", "label": "Value"}],
    )
    return excel_export.stream_workbook(wb, "enternly_dashboard.xlsx")


# ─── Requisitions ─────────────────────────────────────────────────────────────

@router.get("/requisitions/full")
def list_requisitions_full(
    response: Response,
    limit: int = Query(200, ge=1, le=500),
    offset: int = Query(0, ge=0),
    user: dict = Depends(get_current_user),
):
    _deny_hrbp(user)
    role = user["role"]
    uid  = user["sub"]
    if role == "recruiter":
        total = query_one(
            """SELECT COUNT(*) AS n
               FROM requisition r
               JOIN requisition_recruiter rr ON rr.requisition_id = r.id AND rr.recruiter_id = %s
               WHERE COALESCE(r.approval_status,'approved')='approved'""",
            [uid],
        )["n"]
        response.headers["X-Total-Count"] = str(total)
        return query(
            """
            SELECT r.id, r.title, r.status, r.roll_type, r.fiscal_year,
                   r.is_p1, r.risk, r.hiring_location,
                   b.code AS band, bu.name AS business_unit,
                   r.hiring_manager_id,
                   hm.full_name AS hiring_manager_name,
                   (SELECT COUNT(*) FROM application  WHERE requisition_id = r.id) AS in_pipeline,
                   (SELECT COUNT(*) FROM round_config WHERE requisition_id = r.id) AS levels
            FROM requisition r
            JOIN requisition_recruiter rr ON rr.requisition_id = r.id AND rr.recruiter_id = %s
            JOIN band b          ON b.id = r.band_id
            JOIN business_unit bu ON bu.id = r.bu_id
            LEFT JOIN app_user hm ON hm.id = r.hiring_manager_id
            WHERE COALESCE(r.approval_status,'approved')='approved'
            ORDER BY r.created_at DESC
            LIMIT %s OFFSET %s
            """,
            [uid, limit, offset],
        )
    total = query_one(
        """SELECT COUNT(*) AS n FROM requisition r
           WHERE COALESCE(r.approval_status,'approved')='approved'""",
    )["n"]
    response.headers["X-Total-Count"] = str(total)
    return query(
        """
        SELECT r.id, r.title, r.status, r.roll_type, r.fiscal_year,
               r.is_p1, r.risk, r.hiring_location,
               b.code AS band, bu.name AS business_unit,
               r.hiring_manager_id,
               hm.full_name AS hiring_manager_name,
               (SELECT COUNT(*) FROM application  WHERE requisition_id = r.id) AS in_pipeline,
               (SELECT COUNT(*) FROM round_config WHERE requisition_id = r.id) AS levels
        FROM requisition r
        JOIN band b          ON b.id = r.band_id
        JOIN business_unit bu ON bu.id = r.bu_id
        LEFT JOIN app_user hm ON hm.id = r.hiring_manager_id
        WHERE COALESCE(r.approval_status,'approved')='approved'
        ORDER BY r.created_at DESC
        LIMIT %s OFFSET %s
        """,
        [limit, offset],
    )


class RoundIn(BaseModel):
    id: Optional[str] = None  # present for a round that already exists in the DB
    sequence: int
    name: str
    round_type: str = "panel"
    is_auto: bool = False
    panelist_emails: list[str] = []
    feedback_form_id: Optional[str] = None
    meeting_link: Optional[str] = None


def _clean_panelist_emails(emails: list[str]) -> list[str]:
    """Validate + dedupe (case-insensitive) a round's panelist roster, dropping
    anything malformed rather than rejecting the whole requisition save --
    same tolerant pattern main.py's /api/schedule uses for panel_emails."""
    from ..services.email_validation import assert_real_email
    seen, cleaned = set(), []
    for raw in emails or []:
        try:
            e = assert_real_email(raw)
        except ValueError as exc:
            print(f"[pipeline] Dropping invalid panelist email: {exc}")
            continue
        if e.lower() not in seen:
            seen.add(e.lower())
            cleaned.append(e)
    return cleaned


def _validate_feedback_form_id(feedback_form_id: Optional[str], round_type: str = "") -> Optional[str]:
    if feedback_form_id:
        row = query_one(
            "SELECT id FROM feedback_form WHERE id = %s AND is_active = true", [feedback_form_id]
        )
        if not row:
            raise HTTPException(422, "Selected feedback form not found or inactive")
        return feedback_form_id
    # No explicit form chosen -- there's no per-round form-picker UI yet, so
    # every new human round should still land on the standard assessment form
    # rather than silently falling through to the bare 4-field default at
    # submit time. bot_interview rounds are AI-scored from the NexAI
    # transcript, never manually scored, so they're deliberately left
    # unassigned (None) here.
    if (round_type or "").strip().lower() == "bot_interview":
        return None
    default_row = query_one(
        "SELECT id FROM feedback_form WHERE LOWER(name) = LOWER('Interview Assessment Form') AND is_active = true"
    )
    return str(default_row["id"]) if default_row else None


class RequisitionIn(BaseModel):
    title: str
    bu_id: str
    band_id: str
    hrbp_id: Optional[str] = None
    hiring_manager_id: Optional[str] = None
    client_id: Optional[str] = None   # set when hiring on behalf of an external client (RPO/staffing)
    roll_type: str = "on_roll"
    capex_opex: str = "na"
    key_skills: list[str] = []
    min_experience: Optional[float] = None
    max_experience: Optional[float] = None
    budgeted_ctc: Optional[float] = None
    budgeted_fixed: Optional[float] = None
    budgeted_variable: Optional[float] = None
    openings: int = 1
    fiscal_year: Optional[str] = None
    job_description: Optional[str] = None
    is_p1: bool = False
    risk: Optional[str] = None
    hiring_location: Optional[str] = None
    project: Optional[str] = None
    grade_level: Optional[str] = None
    priority: Optional[str] = None
    source_channels: list[str] = []
    screening_questions: list[str] = []
    is_fresher_role: bool = False
    resume_weight: Optional[float] = None
    interview_weight: Optional[float] = None
    rounds: list[RoundIn] = []


@router.post("/requisitions")
def create_requisition(body: RequisitionIn, user: dict = Depends(get_current_user)):
    role = user["role"]
    if role not in ("recruiter", "ta_manager", "admin", "hiring_manager", "hrbp"):
        raise HTTPException(403, "Not authorised to create requisitions")

    missing = []
    if not body.title or not body.title.strip():
        missing.append("Job title")
    if not body.bu_id:
        missing.append("Business unit")
    if not body.band_id:
        missing.append("Band")
    if not body.roll_type:
        missing.append("Roll type")
    if not body.openings or body.openings < 1:
        missing.append("Openings")
    if body.min_experience is None:
        missing.append("Min experience (yrs)")
    if body.max_experience is None:
        missing.append("Max experience (yrs)")
    if body.budgeted_fixed is None:
        missing.append("Budgeted Fixed CTC")
    if body.budgeted_variable is None:
        missing.append("Budgeted Variable CTC")
    if not body.priority and not body.risk:
        missing.append("Priority")
    if not body.hiring_location or not body.hiring_location.strip():
        missing.append("Hiring location")
    if not body.project or not body.project.strip():
        missing.append("Project")
    if not body.job_description or not body.job_description.strip():
        missing.append("Job description")
    if not body.key_skills:
        missing.append("Key skills")
    if not body.rounds:
        missing.append("Panel levels")
    if missing:
        raise HTTPException(422, f"Please fill in required field(s): {', '.join(missing)}.")

    # Derive total CTC from fixed + variable if not explicitly provided
    fixed = body.budgeted_fixed
    variable = body.budgeted_variable
    total_ctc = body.budgeted_ctc
    if total_ctc is None and (fixed is not None or variable is not None):
        total_ctc = (fixed or 0) + (variable or 0)

    # Resolve the HRBP snapshot server-side from hrbp_id -- never trust a
    # client-sent email/name, so a later HRBP change doesn't rewrite history.
    hrbp_id = body.hrbp_id or None
    hrbp_email = hrbp_name = None
    if hrbp_id:
        hrbp_row = query_one(
            "SELECT full_name, email FROM hrbp WHERE id = %s AND is_active = true", [hrbp_id]
        )
        if not hrbp_row:
            raise HTTPException(422, "Selected HRBP not found or inactive")
        hrbp_email, hrbp_name = hrbp_row["email"], hrbp_row["full_name"]

    # Resolve/validate the hiring manager. A hiring_manager creating their own
    # requisition defaults to themselves if none was explicitly picked — every
    # requisition needs an owning HM so the interview self-scheduling flow
    # (scheduling_api._insert_schedule_request_tx) can resolve who to email.
    hiring_manager_id = body.hiring_manager_id or (user["sub"] if role == "hiring_manager" else None)
    if hiring_manager_id:
        hm_row = query_one(
            "SELECT id FROM app_user WHERE id = %s AND role = 'hiring_manager' AND is_active = true",
            [hiring_manager_id],
        )
        if not hm_row:
            raise HTTPException(422, "Selected hiring manager not found or inactive")

    # Resolve/validate the client this req is being worked on behalf of, if
    # any (staffing/RPO provision) -- scoped to the creator's own tenant so
    # one customer can never point a requisition at another's client roster.
    client_id = body.client_id or None
    if client_id:
        client_row = query_one(
            "SELECT id FROM client WHERE id = %s AND tenant_id = %s AND is_active = true",
            [client_id, user.get("tenant_id")],
        )
        if not client_row:
            raise HTTPException(422, "Selected client not found or inactive")

    # Hiring manager reqs start as pending_ta_approval — not visible until TA approves
    approval_status = "pending_ta_approval" if role == "hiring_manager" else "approved"

    rw = body.resume_weight if body.resume_weight is not None else (0.40 if not body.is_fresher_role else 0.40)
    iw = body.interview_weight if body.interview_weight is not None else (0.60 if not body.is_fresher_role else 0.60)

    # Auto-generate req_code, retrying on a concurrent-request collision instead
    # of letting the UNIQUE constraint violation crash the request with a raw
    # 500 and lose the submitted form data.
    import psycopg2
    req = None
    for _attempt in range(5):
        seq_row = query_one(
            "SELECT COALESCE(MAX(CAST(REGEXP_REPLACE(req_code,'[^0-9]','','g') AS INTEGER)),0)+1 AS n FROM requisition WHERE req_code ~ '^REQ-[0-9]+'",
            [],
        )
        req_code = f"REQ-{int((seq_row or {}).get('n') or 1):04d}"
        try:
            req = query_one(
                """
                INSERT INTO requisition
                  (title, bu_id, band_id, hrbp_id, hrbp_email, hrbp_name, hiring_manager_id,
                   client_id, tenant_id,
                   roll_type, capex_opex, key_skills, min_experience, max_experience,
                   budgeted_ctc, budgeted_fixed, budgeted_variable,
                   openings, fiscal_year, job_description,
                   is_p1, risk, hiring_location,
                   project, grade_level, priority, source_channels,
                   screening_questions, is_fresher_role, resume_weight, interview_weight,
                   req_code, status, opened_at, created_by,
                   approval_status, created_by_role)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'open',now(),%s,%s,%s)
                RETURNING id, title, status, req_code, approval_status
                """,
                [
                    body.title, body.bu_id, body.band_id, hrbp_id, hrbp_email, hrbp_name, hiring_manager_id,
                    client_id, user.get("tenant_id"),
                    body.roll_type, body.capex_opex,
                    body.key_skills, body.min_experience, body.max_experience, total_ctc,
                    fixed, variable,
                    body.openings, body.fiscal_year, body.job_description,
                    body.is_p1, body.risk, body.hiring_location,
                    body.project, body.grade_level, body.priority, body.source_channels,
                    body.screening_questions, body.is_fresher_role, rw, iw,
                    req_code, user["sub"], approval_status, role,
                ],
            )
            break
        except psycopg2.errors.UniqueViolation:
            continue
    if req is None:
        raise HTTPException(409, "Could not generate a unique requisition code — please retry.")

    # Auto-assign the creating recruiter as owner -- a ta_manager can also be
    # an individual contributor (personally own/work a requisition), not just
    # a recruiter, so their own analytics/reports aren't left empty.
    if role in ("recruiter", "ta_manager"):
        query(
            """INSERT INTO requisition_recruiter
               (requisition_id, recruiter_id, is_owner, assigned_by)
               VALUES (%s,%s,true,%s)""",
            [req["id"], user["sub"], user["sub"]],
            fetch=False,
        )

    # Notify TA managers when a hiring manager creates a requisition needing approval
    if role == "hiring_manager":
        hm_row = query_one("SELECT full_name FROM app_user WHERE id=%s", [user["sub"]])
        hm_name = (hm_row or {}).get("full_name") or "Hiring Manager"
        ta_emails = query(
            "SELECT email FROM app_user WHERE role='ta_manager' AND is_active=TRUE", []
        )
        ta_email_list = [r["email"] for r in (ta_emails or []) if r.get("email")]
        if ta_email_list:
            try:
                from ..services.email_templates import render_template as _rt
                from ..services.connectors import send_email as _se, resolve_global_placeholders
                from ..services.email_layout import build_branded_email
                req_id_for_globals = str(req["id"]) if req else None
                globals_ = resolve_global_placeholders(req_id=req_id_for_globals, actor=user)
                reply_to = globals_.get("recruiter_email") or None
                subj, bdy = _rt("hm_req_approval_request", {
                    "hm_name":   hm_name,
                    "req_title": body.title,
                }, req_id=req_id_for_globals, actor=user)
                html_body = build_branded_email(
                    eyebrow="Application Tracking System",
                    hero_title_html="Approval<br>Needed.",
                    hero_subtitle="A hiring manager has submitted a new requisition for your review.",
                    detail_cells=[("Requisition", body.title), ("Submitted By", hm_name)],
                    about_text=bdy,
                    about_heading=None,
                    cta_label=None, cta_link=None,
                )
                for addr in ta_email_list:
                    try:
                        _se(addr, subj, bdy, html=html_body, reply_to=reply_to)
                    except Exception as exc:
                        print(f"[pipeline] ta-notification to {addr} failed: {exc}")
            except Exception as exc:
                print(f"[pipeline] hm req notification failed: {exc}")

    # Create panel rounds
    for r in body.rounds:
        query(
            """INSERT INTO round_config
               (requisition_id, sequence, name, round_type, is_auto, panelist_emails, feedback_form_id, meeting_link)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
            [req["id"], r.sequence, r.name, r.round_type, r.is_auto,
             _clean_panelist_emails(r.panelist_emails), _validate_feedback_form_id(r.feedback_form_id, r.round_type),
             (r.meeting_link or None)],
            fetch=False,
        )

    log_activity(
        "requisition", "requisition_created",
        entity_id=req["id"], requisition_id=req["id"],
        actor_id=user["sub"], actor_role=role,
        to_value=req.get("status"),
        detail={"title": body.title, "req_code": req.get("req_code"), "approval_status": approval_status},
    )
    return req


_JD_STORE = _os.environ.get("JD_STORE_DIR", "/app/jd_store")
_JD_SUPPORTED_EXT = ("pdf", "docx", "doc")

_JD_PARSE_SYSTEM = """\
You are a job description parser. Extract structured data from the job description text.
Return ONLY a valid JSON object — no markdown fences, no prose before or after.
Do NOT rewrite or summarize the job description — the caller keeps the original text verbatim.

Required fields:
{
  "job_title": "<role title, or null>",
  "key_skills": ["list", "of", "required", "skills"],
  "min_experience": <minimum years of experience as a number, or null>,
  "max_experience": <maximum years of experience as a number, or null>,
  "hiring_location": "<city or location mentioned, or null>",
  "band_or_grade": "<band, grade, or level if mentioned, or null>"
}"""


def _jd_safe_float(val) -> Optional[float]:
    try:
        return float(val) if val is not None else None
    except (TypeError, ValueError):
        return None


@router.post("/requisitions/parse-jd")
async def parse_jd(
    file: Optional[UploadFile] = File(None),
    raw_text: str = Form(""),
    user: dict = Depends(get_current_user),
):
    role = user.get("role", "")
    if role not in ("recruiter", "ta_manager", "hiring_manager", "admin", "hrbp"):
        raise HTTPException(403, "Not authorised to use JD parsing")

    # Extract text from file or use pasted text
    text = ""
    if file and file.filename:
        suffix = _os.path.splitext(file.filename or "")[1].lower().lstrip(".")
        if suffix not in ("pdf", "docx", "doc"):
            return JSONResponse(
                status_code=422,
                content={"detail": "Unsupported file type. Upload a PDF or Word document (.pdf, .docx, .doc)."},
            )
        file_bytes = await file.read()
        from ..services.cv_parser import extract_text as _cv_extract_text
        text = _cv_extract_text(file_bytes, suffix)
        if not text.strip():
            return JSONResponse(
                status_code=422,
                content={"detail": "Could not extract text from the uploaded file. Try pasting the JD text instead."},
            )
    elif raw_text.strip():
        text = raw_text.strip()
    else:
        return JSONResponse(
            status_code=422,
            content={"detail": "Provide a JD file or paste JD text."},
        )

    # Call Groq — reuse the shared async client/model from interviewer_llm
    # (same singleton, same key source, same env-var validation as NexAI/screening)
    try:
        from ..services.interviewer_llm import _get_client as _groq_client, _model as _groq_model
        client = _groq_client()
        resp = await client.chat.completions.create(
            model=_groq_model(),
            messages=[
                {"role": "system", "content": _JD_PARSE_SYSTEM},
                {"role": "user", "content": f"Job description:\n\n{text[:6000]}"},
            ],
            temperature=0,
            max_tokens=600,
        )
        raw_resp = resp.choices[0].message.content or ""
        # Strip markdown fences before parsing
        cleaned = _re.sub(r"^```(?:json)?\s*", "", raw_resp.strip(), flags=_re.IGNORECASE)
        cleaned = _re.sub(r"\s*```$", "", cleaned.strip())
        data = _json.loads(cleaned)
    except _json.JSONDecodeError:
        return JSONResponse(
            status_code=422,
            content={"detail": "The AI could not parse this JD into structured data. Try pasting cleaner text."},
        )
    except Exception as exc:
        return JSONResponse(
            status_code=422,
            content={"detail": f"JD parsing failed: {str(exc)[:120]}"},
        )

    # Normalise key_skills — could be a string or a list
    skills = data.get("key_skills") or []
    if isinstance(skills, str):
        skills = [s.strip() for s in skills.split(",") if s.strip()]

    return {
        "job_title":       data.get("job_title") or None,
        "key_skills":      skills,
        "min_experience":  _jd_safe_float(data.get("min_experience")),
        "max_experience":  _jd_safe_float(data.get("max_experience")),
        # The full original text, verbatim — not the LLM's summary, so long/rich
        # JDs never get silently condensed down to a short paragraph.
        "job_description": text.strip(),
        "hiring_location": data.get("hiring_location") or None,
        "band_or_grade":   data.get("band_or_grade") or None,
    }


@router.post("/requisitions/{req_id}/jd-file")
async def upload_jd_file(
    req_id: str,
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """Attach the original JD file to a requisition so it can be downloaded
    later — separate from parse-jd, which only extracts text and never
    persisted the file itself."""
    role = user.get("role", "")
    if role not in ("recruiter", "ta_manager", "hiring_manager", "admin", "hrbp"):
        raise HTTPException(403, "Not authorised to attach a JD file")

    req = query_one("SELECT id FROM requisition WHERE id = %s", [req_id])
    if not req:
        raise HTTPException(404, "requisition not found")

    if role == "recruiter" and not query_one(
        "SELECT 1 FROM requisition_recruiter WHERE requisition_id=%s AND recruiter_id=%s",
        [req_id, user["sub"]],
    ):
        raise HTTPException(403, "Not authorised")

    suffix = _os.path.splitext(file.filename or "")[1].lower().lstrip(".")
    if suffix not in _JD_SUPPORTED_EXT:
        raise HTTPException(422, "Unsupported file type. Upload a PDF or Word document (.pdf, .docx, .doc).")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(422, "Uploaded file is empty.")

    _os.makedirs(_JD_STORE, exist_ok=True)
    dest = _os.path.join(_JD_STORE, f"{req_id}.{suffix}")
    with open(dest, "wb") as f:
        f.write(file_bytes)

    query(
        "UPDATE requisition SET jd_file_path=%s, jd_file_name=%s WHERE id=%s",
        [dest, file.filename, req_id],
        fetch=False,
    )
    return {"jd_file_name": file.filename}


@router.get("/requisitions/{req_id}/jd-file")
def download_jd_file(req_id: str, user: dict = Depends(get_current_user)):
    _deny_hrbp(user)
    req = query_one(
        "SELECT jd_file_path, jd_file_name FROM requisition WHERE id = %s", [req_id]
    )
    if not req or not req.get("jd_file_path"):
        raise HTTPException(404, "No JD file attached to this requisition")
    if not _os.path.isfile(req["jd_file_path"]):
        raise HTTPException(404, "JD file is missing from storage")
    return FileResponse(
        req["jd_file_path"],
        filename=req.get("jd_file_name") or "job_description",
        media_type="application/octet-stream",
    )


@router.get("/requisitions/{req_id}/detail")
def get_requisition_detail(req_id: str, user: dict = Depends(get_current_user)):
    _deny_hrbp(user)
    req = query_one(
        """
        SELECT r.*, b.code AS band_code, bu.name AS business_unit_name
        FROM requisition r
        JOIN band b          ON b.id = r.band_id
        JOIN business_unit bu ON bu.id = r.bu_id
        WHERE r.id = %s
        """,
        [req_id],
    )
    if not req:
        raise HTTPException(404, "requisition not found")
    rounds = query(
        "SELECT * FROM round_config WHERE requisition_id = %s ORDER BY sequence",
        [req_id],
    )
    return {**dict(req), "rounds": rounds}


# ─── Edit requisition ─────────────────────────────────────────────────────────

class RequisitionEditIn(BaseModel):
    title: Optional[str] = None
    bu_id: Optional[str] = None
    band_id: Optional[str] = None
    hrbp_id: Optional[str] = None
    hiring_manager_id: Optional[str] = None
    client_id: Optional[str] = None   # "" clears it back to an internal hire
    roll_type: Optional[str] = None
    capex_opex: Optional[str] = None
    key_skills: Optional[list[str]] = None
    min_experience: Optional[float] = None
    max_experience: Optional[float] = None
    budgeted_fixed: Optional[float] = None
    budgeted_variable: Optional[float] = None
    openings: Optional[int] = None
    fiscal_year: Optional[str] = None
    job_description: Optional[str] = None
    is_p1: Optional[bool] = None
    risk: Optional[str] = None
    hiring_location: Optional[str] = None
    project: Optional[str] = None
    grade_level: Optional[str] = None
    priority: Optional[str] = None
    source_channels: Optional[list[str]] = None
    screening_questions: Optional[list[str]] = None
    is_fresher_role: Optional[bool] = None
    resume_weight: Optional[float] = None
    interview_weight: Optional[float] = None
    rounds: Optional[list[RoundIn]] = None


@router.patch("/requisitions/{req_id}")
def edit_requisition(req_id: str, body: RequisitionEditIn, user: dict = Depends(get_current_user)):
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised to edit requisitions")

    existing = query_one("SELECT id, status FROM requisition WHERE id = %s", [req_id])
    if not existing:
        raise HTTPException(404, "Requisition not found")
    if user["role"] == "recruiter" and not _recruiter_owns_req(user, req_id):
        raise HTTPException(404, "Requisition not found")

    sets, vals = [], []

    def _add(col, val):
        sets.append(f"{col} = %s")
        vals.append(val)

    if body.title is not None:           _add("title", body.title)
    if body.bu_id is not None:           _add("bu_id", body.bu_id)
    if body.band_id is not None:         _add("band_id", body.band_id)
    if body.hrbp_id is not None:
        if body.hrbp_id == "":
            _add("hrbp_id", None); _add("hrbp_email", None); _add("hrbp_name", None)
        else:
            hrbp_row = query_one(
                "SELECT full_name, email FROM hrbp WHERE id = %s AND is_active = true", [body.hrbp_id]
            )
            if not hrbp_row:
                raise HTTPException(422, "Selected HRBP not found or inactive")
            _add("hrbp_id", body.hrbp_id)
            _add("hrbp_email", hrbp_row["email"])
            _add("hrbp_name", hrbp_row["full_name"])
    if body.hiring_manager_id is not None:
        if body.hiring_manager_id == "":
            _add("hiring_manager_id", None)
        else:
            hm_row = query_one(
                "SELECT id FROM app_user WHERE id = %s AND role = 'hiring_manager' AND is_active = true",
                [body.hiring_manager_id],
            )
            if not hm_row:
                raise HTTPException(422, "Selected hiring manager not found or inactive")
            _add("hiring_manager_id", body.hiring_manager_id)
    if body.client_id is not None:
        if body.client_id == "":
            _add("client_id", None)
        else:
            client_row = query_one(
                "SELECT id FROM client WHERE id = %s AND tenant_id = %s AND is_active = true",
                [body.client_id, user.get("tenant_id")],
            )
            if not client_row:
                raise HTTPException(422, "Selected client not found or inactive")
            _add("client_id", body.client_id)
    if body.roll_type is not None:       _add("roll_type", body.roll_type)
    if body.capex_opex is not None:      _add("capex_opex", body.capex_opex)
    if body.key_skills is not None:      _add("key_skills", body.key_skills)
    if body.min_experience is not None:  _add("min_experience", body.min_experience)
    if body.max_experience is not None:  _add("max_experience", body.max_experience)
    if body.openings is not None:        _add("openings", body.openings)
    if body.fiscal_year is not None:     _add("fiscal_year", body.fiscal_year)
    if body.job_description is not None: _add("job_description", body.job_description)
    if body.is_p1 is not None:           _add("is_p1", body.is_p1)
    if body.risk is not None:            _add("risk", body.risk or None)
    if body.hiring_location is not None: _add("hiring_location", body.hiring_location or None)
    if body.project is not None:         _add("project", body.project or None)
    if body.grade_level is not None:     _add("grade_level", body.grade_level or None)
    if body.priority is not None:        _add("priority", body.priority or None)
    if body.source_channels is not None: _add("source_channels", body.source_channels)
    if body.screening_questions is not None: _add("screening_questions", body.screening_questions)
    if body.is_fresher_role is not None: _add("is_fresher_role", body.is_fresher_role)
    if body.resume_weight is not None:   _add("resume_weight", body.resume_weight)
    if body.interview_weight is not None: _add("interview_weight", body.interview_weight)

    if body.budgeted_fixed is not None or body.budgeted_variable is not None:
        fixed = body.budgeted_fixed
        variable = body.budgeted_variable
        if fixed is not None:    _add("budgeted_fixed", fixed)
        if variable is not None: _add("budgeted_variable", variable)
        # Recalculate total only when both sides known
        existing_row = query_one(
            "SELECT budgeted_fixed, budgeted_variable FROM requisition WHERE id=%s", [req_id]
        )
        f = fixed if fixed is not None else (existing_row or {}).get("budgeted_fixed") or 0
        v = variable if variable is not None else (existing_row or {}).get("budgeted_variable") or 0
        _add("budgeted_ctc", f + v)

    changed_fields = [s.split(" = ")[0] for s in sets]
    if sets:
        vals.append(req_id)
        query(f"UPDATE requisition SET {', '.join(sets)} WHERE id = %s", vals, fetch=False)

    # Reconcile round_config against the submitted list. Existing rounds are
    # matched by id and UPDATEd in place (never dropped + recreated) so that
    # interview/scorecard/scheduling rows referencing round_config.id stay
    # valid — a requisition can freely have rounds edited, reordered, or have
    # new rounds inserted at any position (start/middle/end) even after
    # candidates already have interviews recorded against existing rounds.
    if body.rounds is not None:
        existing_ids = {
            str(row["id"])
            for row in query("SELECT id FROM round_config WHERE requisition_id = %s", [req_id])
        }
        submitted_ids = {r.id for r in body.rounds if r.id and r.id in existing_ids}
        removed_ids = existing_ids - submitted_ids

        if removed_ids:
            blocked = query(
                """SELECT rc.id, rc.name FROM round_config rc
                   WHERE rc.id = ANY(%s)
                     AND EXISTS (SELECT 1 FROM interview iv WHERE iv.round_config_id = rc.id)""",
                [list(removed_ids)],
            )
            if blocked:
                names = ", ".join(f'"{b["name"]}"' for b in blocked)
                raise HTTPException(
                    422,
                    f"Can't remove round(s) {names} — candidate interview(s) already exist against "
                    "them. Rename/reconfigure the round instead, or leave it in place.",
                )

        # Validate everything up front so a bad feedback_form_id can't leave
        # round_config half-migrated.
        prepared = [
            (r, _validate_feedback_form_id(r.feedback_form_id, r.round_type), _clean_panelist_emails(r.panelist_emails))
            for r in body.rounds
        ]

        # Bump existing rows out of the active 1..N sequence range first so
        # reordering / mid-list inserts below never collide with the
        # UNIQUE(requisition_id, sequence) constraint mid-update.
        query("UPDATE round_config SET sequence = -sequence - 1 WHERE requisition_id = %s", [req_id], fetch=False)

        for r, feedback_form_id, panelist_emails in prepared:
            if r.id and r.id in existing_ids:
                query(
                    """UPDATE round_config
                       SET sequence = %s, name = %s, round_type = %s, is_auto = %s,
                           panelist_emails = %s, feedback_form_id = %s, meeting_link = %s
                       WHERE id = %s AND requisition_id = %s""",
                    [r.sequence, r.name, r.round_type, r.is_auto,
                     panelist_emails, feedback_form_id, (r.meeting_link or None), r.id, req_id],
                    fetch=False,
                )
            else:
                query(
                    """INSERT INTO round_config (requisition_id, sequence, name, round_type, is_auto, panelist_emails, feedback_form_id, meeting_link)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                    [req_id, r.sequence, r.name, r.round_type, r.is_auto,
                     panelist_emails, feedback_form_id, (r.meeting_link or None)],
                    fetch=False,
                )

        if removed_ids:
            query("DELETE FROM round_config WHERE id = ANY(%s)", [list(removed_ids)], fetch=False)

        changed_fields.append("rounds")

    if changed_fields:
        log_activity(
            "requisition", "requisition_updated",
            entity_id=req_id, requisition_id=req_id,
            actor_id=user["sub"], actor_role=user["role"],
            detail={"changed_fields": changed_fields},
        )

    return {"ok": True}


# ─── Status change (hold / cancel / reopen) ───────────────────────────────────

class ReqStatusIn(BaseModel):
    status: str


@router.patch("/requisitions/{req_id}/status")
def change_req_status(req_id: str, body: ReqStatusIn, user: dict = Depends(get_current_user)):
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised to change requisition status")
    allowed = {"open", "on_hold", "closed", "cancelled"}
    if body.status not in allowed:
        raise HTTPException(400, f"Status must be one of: {', '.join(allowed)}")
    existing = query_one("SELECT id, status FROM requisition WHERE id = %s", [req_id])
    if not existing:
        raise HTTPException(404, "Requisition not found")
    if user["role"] == "recruiter" and not _recruiter_owns_req(user, req_id):
        raise HTTPException(404, "Requisition not found")

    extra = {}
    if body.status == "open" and existing["status"] != "open":
        extra["opened_at"] = "now()"
        extra["closed_at"] = "clear"
    if body.status in ("closed", "cancelled"):
        extra["closed_at"] = "now()"

    set_clause = "status = %s"
    vals = [body.status]
    if "opened_at" in extra:
        set_clause += ", opened_at = now()"
    if extra.get("closed_at") == "now()":
        set_clause += ", closed_at = now()"
    elif extra.get("closed_at") == "clear":
        set_clause += ", closed_at = NULL"
    vals.append(req_id)

    query(f"UPDATE requisition SET {set_clause} WHERE id = %s", vals, fetch=False)
    log_activity(
        "requisition", "requisition_status_changed",
        entity_id=req_id, requisition_id=req_id,
        actor_id=user["sub"], actor_role=user["role"],
        from_value=existing["status"], to_value=body.status,
    )
    return {"ok": True, "status": body.status}


@router.get("/requisitions/{req_id}/kanban")
def kanban(req_id: str, user: dict = Depends(get_current_user)):
    _deny_hrbp(user)
    # Role scope: recruiter must own the req (same guard as get_req_pipeline)
    if user["role"] == "recruiter":
        if not query_one(
            "SELECT 1 FROM requisition_recruiter WHERE requisition_id=%s AND recruiter_id=%s",
            [req_id, user["sub"]],
        ):
            raise HTTPException(403, "Not authorised")
    rounds = query(
        """SELECT id, sequence, name, round_type, is_auto, panelist_emails
           FROM round_config WHERE requisition_id = %s ORDER BY sequence""",
        [req_id],
    )
    candidates = query(
        """
        SELECT a.id AS app_id, a.status, a.current_round, a.flags,
               COALESCE(a.combined_score, a.match_score) AS score,
               (a.combined_score IS NOT NULL) AS nexai_completed,
               c.full_name, c.gender, c.email,
               iv.id AS interview_id, iv.scheduled_at AS interview_scheduled_at, iv.status AS interview_status
        FROM application a
        JOIN candidate c ON c.id = a.candidate_id
        LEFT JOIN LATERAL (
            SELECT i.id, i.scheduled_at, i.status
            FROM interview i
            JOIN round_config rc ON rc.id = i.round_config_id
            WHERE i.application_id = a.id AND rc.sequence = a.current_round
            ORDER BY i.created_at DESC LIMIT 1
        ) iv ON true
        WHERE a.requisition_id = %s
        ORDER BY score DESC NULLS LAST
        """,
        [req_id],
    )
    return {"rounds": rounds, "candidates": candidates}


# ─── Recruiter Assignment ─────────────────────────────────────────────────────

@router.get("/requisitions/{req_id}/recruiters")
def get_req_recruiters(req_id: str, user: dict = Depends(get_current_user)):
    if user["role"] not in ("ta_manager", "admin", "recruiter"):
        raise HTTPException(403, "Not authorised")
    if user["role"] == "recruiter" and not _recruiter_owns_req(user, req_id):
        raise HTTPException(404, "Not found")
    return query(
        """SELECT u.id, u.full_name, u.email, rr.is_owner, rr.assigned_at
           FROM requisition_recruiter rr
           JOIN app_user u ON u.id = rr.recruiter_id
           WHERE rr.requisition_id = %s
           ORDER BY rr.is_owner DESC, rr.assigned_at""",
        [req_id],
    )


class AssignRecruiterIn(BaseModel):
    recruiter_id: str


@router.post("/requisitions/{req_id}/assign-recruiter")
def assign_recruiter(req_id: str, body: AssignRecruiterIn, user: dict = Depends(get_current_user)):
    if user["role"] not in ("ta_manager", "admin"):
        raise HTTPException(403, "Only TA managers can assign recruiters")
    req = query_one("SELECT id FROM requisition WHERE id = %s", [req_id])
    if not req:
        raise HTTPException(404, "Requisition not found")
    # A ta_manager can also be assigned as a working recruiter on a
    # requisition (they can be an individual contributor too, not just a
    # manager) -- so their own analytics/reports reflect requisitions they
    # personally work, same as any recruiter.
    recruiter = query_one(
        "SELECT id FROM app_user WHERE id = %s AND role IN ('recruiter','ta_manager') AND is_active = true",
        [body.recruiter_id],
    )
    if not recruiter:
        raise HTTPException(404, "Active recruiter not found")
    query(
        """INSERT INTO requisition_recruiter (requisition_id, recruiter_id, is_owner, assigned_by)
           VALUES (%s, %s, false, %s)
           ON CONFLICT (requisition_id, recruiter_id) DO NOTHING""",
        [req_id, body.recruiter_id, user["sub"]],
        fetch=False,
    )
    log_activity(
        "requisition", "recruiter_assigned",
        entity_id=req_id, requisition_id=req_id,
        actor_id=user["sub"], actor_role=user["role"],
        to_value=body.recruiter_id,
    )
    return {"ok": True}


@router.delete("/requisitions/{req_id}/recruiters/{recruiter_id}")
def unassign_recruiter(req_id: str, recruiter_id: str, user: dict = Depends(get_current_user)):
    if user["role"] not in ("ta_manager", "admin"):
        raise HTTPException(403, "Only TA managers can remove assignments")
    query(
        "DELETE FROM requisition_recruiter WHERE requisition_id = %s AND recruiter_id = %s",
        [req_id, recruiter_id],
        fetch=False,
    )
    return {"ok": True}


# ─── Team ──────────────────────────────────────────────────────────────────────

@router.get("/team")
def get_team(user: dict = Depends(get_current_user)):
    if user["role"] not in ("ta_manager", "admin"):
        raise HTTPException(403, "TA Manager access required")
    return query(
        """
        SELECT u.id, u.full_name, u.email, u.role,
               COUNT(DISTINCT rr.requisition_id)
                 FILTER (WHERE r.status = 'open') AS active_req_count,
               COALESCE(
                 json_agg(
                   json_build_object('req_id', r.id, 'title', r.title, 'status', r.status)
                 ) FILTER (WHERE r.id IS NOT NULL AND r.status = 'open'),
                 '[]'::json
               ) AS assigned_requisitions
        FROM app_user u
        LEFT JOIN requisition_recruiter rr ON rr.recruiter_id = u.id
        LEFT JOIN requisition r ON r.id = rr.requisition_id
        WHERE u.role IN ('recruiter', 'ta_manager', 'hiring_manager')
          AND u.is_active = true
        GROUP BY u.id, u.full_name, u.email, u.role
        ORDER BY u.role, u.full_name
        """
    )


# ─── Candidates ───────────────────────────────────────────────────────────────

@router.get("/candidates")
def list_candidates(
    response: Response,
    user: dict = Depends(get_current_user),
    q: Optional[str] = None,
    limit: int = Query(200, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    _deny_hrbp(user)
    role = user["role"]
    uid  = user["sub"]
    # Recruiter LATERAL sub-select: owner recruiter of each requisition
    _rec_lat = """
        LEFT JOIN LATERAL (
            SELECT rr2.recruiter_id, ru.full_name AS recruiter_name
            FROM requisition_recruiter rr2
            JOIN app_user ru ON ru.id = rr2.recruiter_id
            WHERE rr2.requisition_id = r.id
            ORDER BY rr2.is_owner DESC NULLS LAST
            LIMIT 1
        ) rc_info ON true
    """
    # Search: name / email / phone (ILIKE) or skills pulled from the linked CV Repository entry.
    _search_where = ""
    _search_params: list = []
    if q and q.strip():
        term = f"%{q.strip()}%"
        _search_where = """
            AND (c.full_name ILIKE %s OR c.email ILIKE %s OR c.phone ILIKE %s
                 OR EXISTS (
                     SELECT 1 FROM cv_repository cvr
                     WHERE cvr.candidate_id = c.id
                       AND EXISTS (SELECT 1 FROM unnest(cvr.skills) sk WHERE sk ILIKE %s)
                 ))
        """
        _search_params = [term, term, term, term]
    from ..services.source_labels import attach_source_labels

    if role == "recruiter":
        total = query_one(
            f"""
            SELECT COUNT(*) AS n FROM (
              SELECT DISTINCT ON (LOWER(c.email), r.id) 1
              FROM candidate c
              JOIN application  a ON a.candidate_id = c.id
              JOIN requisition  r ON r.id = a.requisition_id
              JOIN requisition_recruiter rr
                   ON rr.requisition_id = r.id AND rr.recruiter_id = %s
              WHERE 1=1 {_search_where}
            ) deduped
            """,
            [uid] + _search_params,
        )["n"]
        response.headers["X-Total-Count"] = str(total)
        rows = query(
            f"""
            SELECT * FROM (
              SELECT DISTINCT ON (LOWER(c.email), r.id)
                c.id, c.full_name, c.email, c.gender,
                r.id AS req_id, r.title AS requisition, a.status, a.flags,
                a.combined_score, a.match_score, a.id AS app_id,
                rc_info.recruiter_id, rc_info.recruiter_name,
                a.source AS app_source,
                a.applied_at
              FROM candidate c
              JOIN application  a ON a.candidate_id = c.id
              JOIN requisition  r ON r.id = a.requisition_id
              JOIN requisition_recruiter rr
                   ON rr.requisition_id = r.id AND rr.recruiter_id = %s
              {_rec_lat}
              WHERE 1=1 {_search_where}
              ORDER BY LOWER(c.email), r.id, a.combined_score DESC NULLS LAST, a.applied_at DESC
            ) deduped
            ORDER BY applied_at DESC
            LIMIT %s OFFSET %s
            """,
            [uid] + _search_params + [limit, offset],
        ) or []
        attach_source_labels(rows, "app_source")
        return rows

    total = query_one(
        f"""
        SELECT COUNT(*) AS n FROM (
          SELECT DISTINCT ON (LOWER(c.email), r.id) 1
          FROM candidate c
          JOIN application a ON a.candidate_id = c.id
          JOIN requisition r ON r.id = a.requisition_id
          WHERE 1=1 {_search_where}
        ) deduped
        """,
        _search_params,
    )["n"]
    response.headers["X-Total-Count"] = str(total)
    rows = query(
        f"""
        SELECT * FROM (
          SELECT DISTINCT ON (LOWER(c.email), r.id)
            c.id, c.full_name, c.email, c.gender,
            r.id AS req_id, r.title AS requisition, a.status, a.flags,
            a.combined_score, a.match_score, a.id AS app_id,
            rc_info.recruiter_id, rc_info.recruiter_name,
            a.source AS app_source,
            a.applied_at
          FROM candidate c
          JOIN application a ON a.candidate_id = c.id
          JOIN requisition r ON r.id = a.requisition_id
          {_rec_lat}
          WHERE 1=1 {_search_where}
          ORDER BY LOWER(c.email), r.id, a.combined_score DESC NULLS LAST, a.applied_at DESC
        ) deduped
        ORDER BY applied_at DESC
        LIMIT %s OFFSET %s
        """,
        _search_params + [limit, offset],
    ) or []
    attach_source_labels(rows, "app_source")
    return rows


@router.get("/candidates/{candidate_id}")
def get_candidate_detail(candidate_id: str, user: dict = Depends(get_current_user)):
    """Full candidate profile: contact info, every application, and the linked CV/skills."""
    _deny_hrbp(user)
    role = user["role"]
    uid  = user["sub"]

    cand = query_one(
        """SELECT id, full_name, email, phone, gender, source, resume_url, created_at
           FROM candidate WHERE id = %s""",
        [candidate_id],
    )
    if not cand:
        raise HTTPException(404, "Candidate not found")

    if role == "recruiter":
        owns = query_one(
            """SELECT 1 FROM application a
               JOIN requisition_recruiter rr ON rr.requisition_id = a.requisition_id
               WHERE a.candidate_id = %s AND rr.recruiter_id = %s LIMIT 1""",
            [candidate_id, uid],
        )
        if not owns:
            raise HTTPException(404, "Candidate not found")

    apps = query(
        """
        SELECT a.id AS app_id, a.requisition_id, r.title AS requisition_title,
               r.req_code, a.status, a.match_score, a.combined_score,
               a.score_breakdown, a.applied_at, a.panel_consensus,
               a.current_company, a.current_designation, a.current_location,
               a.current_ctc_fixed, a.current_ctc_variable, a.current_ctc_bonus, a.current_ctc_total,
               a.expected_ctc_fixed, a.expected_ctc_variable, a.expected_ctc_bonus, a.expected_ctc_total,
               a.notice_period_days, a.willing_to_relocate,
               rc_info.recruiter_name
        FROM application a
        JOIN requisition r ON r.id = a.requisition_id
        LEFT JOIN LATERAL (
            SELECT ru.full_name AS recruiter_name
            FROM requisition_recruiter rr2
            JOIN app_user ru ON ru.id = rr2.recruiter_id
            WHERE rr2.requisition_id = r.id
            ORDER BY rr2.is_owner DESC NULLS LAST
            LIMIT 1
        ) rc_info ON true
        WHERE a.candidate_id = %s
        ORDER BY a.applied_at DESC
        """,
        [candidate_id],
    ) or []

    cv_row = query_one(
        """SELECT id, file_name, skills, experience_years, current_position,
                  location, ai_summary, enrich_status, map_status
           FROM cv_repository
           WHERE candidate_id = %s
           ORDER BY created_at DESC
           LIMIT 1""",
        [candidate_id],
    )
    cv = None
    if cv_row:
        cv = {
            "id":               str(cv_row["id"]),
            "file_name":        cv_row["file_name"],
            "skills":           list(cv_row["skills"] or []),
            "experience_years": cv_row["experience_years"],
            "current_position": cv_row["current_position"],
            "location":         cv_row["location"],
            "ai_summary":       cv_row["ai_summary"],
            "enrich_status":    cv_row["enrich_status"],
            "map_status":       cv_row["map_status"],
        }

    return {
        "id":         str(cand["id"]),
        "full_name":  cand["full_name"],
        "email":      cand["email"],
        "phone":      cand["phone"],
        "gender":     cand["gender"],
        "source":     cand["source"],
        "created_at": cand["created_at"].isoformat() if cand["created_at"] else None,
        "cv":         cv,
        # Fallback link when cv_repository has no linked row (e.g. background
        # ingestion failed) but the raw uploaded file still exists on disk.
        "resume_url": cand["resume_url"] if not cv else None,
        "applications": [
            {
                "app_id":            str(a["app_id"]),
                "requisition_id":    str(a["requisition_id"]) if a["requisition_id"] else None,
                "requisition_title": a["requisition_title"],
                "req_code":          a["req_code"],
                "status":            a["status"],
                "match_score":       a["match_score"],
                "combined_score":    a["combined_score"],
                "score_breakdown":   a["score_breakdown"],
                "applied_at":        a["applied_at"].isoformat() if a["applied_at"] else None,
                "panel_consensus":   a["panel_consensus"],
                "recruiter_name":    a["recruiter_name"],
                "current_company":       a["current_company"],
                "current_designation":   a["current_designation"],
                "current_location":      a["current_location"],
                "current_ctc_fixed":     a["current_ctc_fixed"],
                "current_ctc_variable":  a["current_ctc_variable"],
                "current_ctc_bonus":     a["current_ctc_bonus"],
                "current_ctc_total":     a["current_ctc_total"],
                "expected_ctc_fixed":    a["expected_ctc_fixed"],
                "expected_ctc_variable": a["expected_ctc_variable"],
                "expected_ctc_bonus":    a["expected_ctc_bonus"],
                "expected_ctc_total":    a["expected_ctc_total"],
                "notice_period_days":    a["notice_period_days"],
                "willing_to_relocate":   a["willing_to_relocate"],
            }
            for a in apps
        ],
    }


# ─── Interviews ───────────────────────────────────────────────────────────────

@router.get("/interviews")
def list_interviews(user: dict = Depends(get_current_user)):
    role = user["role"]
    uid  = user["sub"]

    _sc_status_sub = (
        "(SELECT s.status FROM scorecard s "
        "WHERE s.interview_id = i.id AND s.interviewer_id = %s LIMIT 1) AS my_scorecard_status"
    )
    _panel_sub = (
        "EXISTS(SELECT 1 FROM interview_panel ip "
        "WHERE ip.interview_id = i.id AND ip.interviewer_id = %s) AS is_on_panel"
    )

    _notes_sub = (
        "COALESCE((SELECT inotes.fetch_status FROM interview_notes inotes "
        "WHERE inotes.interview_id = i.id LIMIT 1), 'none') AS transcript_status"
    )

    if role == "recruiter":
        return query(
            f"""
            SELECT i.id, i.scheduled_at, i.status, i.meet_link, i.mode,
                   i.gcal_event_id,
                   c.full_name AS candidate_name, r.title AS requisition,
                   rc.name AS round_name, a.id AS application_id,
                   {_panel_sub},
                   {_sc_status_sub},
                   {_notes_sub}
            FROM interview i
            JOIN application  a  ON a.id  = i.application_id
            JOIN candidate    c  ON c.id  = a.candidate_id
            JOIN requisition  r  ON r.id  = a.requisition_id
            JOIN round_config rc ON rc.id = i.round_config_id
            JOIN requisition_recruiter rr
                 ON rr.requisition_id = r.id AND rr.recruiter_id = %s
            ORDER BY i.scheduled_at DESC NULLS LAST
            """,
            [uid, uid, uid],
        )

    if role == "interviewer":
        # Pure interviewers see only their own panel assignments
        return query(
            f"""
            SELECT i.id, i.scheduled_at, i.status, i.meet_link, i.mode,
                   i.gcal_event_id,
                   c.full_name AS candidate_name, r.title AS requisition,
                   rc.name AS round_name, a.id AS application_id,
                   TRUE AS is_on_panel,
                   {_sc_status_sub},
                   {_notes_sub}
            FROM interview i
            JOIN application  a  ON a.id  = i.application_id
            JOIN candidate    c  ON c.id  = a.candidate_id
            JOIN requisition  r  ON r.id  = a.requisition_id
            JOIN round_config rc ON rc.id = i.round_config_id
            JOIN interview_panel ip ON ip.interview_id = i.id AND ip.interviewer_id = %s
            ORDER BY i.scheduled_at DESC NULLS LAST
            """,
            [uid, uid],
        )

    if role == "hrbp":
        # HRBP has no scorecard/panel role on an interview -- just their own
        # requisition scope (same rule as their dashboard/requisitions list),
        # so opening this screen to them doesn't leak every interview company-wide.
        from .hrbp_api import scope_requisitions_for_hrbp
        where, params = scope_requisitions_for_hrbp(user)
        return query(
            f"""
            SELECT i.id, i.scheduled_at, i.status, i.meet_link, i.mode,
                   i.gcal_event_id,
                   c.full_name AS candidate_name, r.title AS requisition,
                   rc.name AS round_name, a.id AS application_id,
                   FALSE AS is_on_panel,
                   NULL AS my_scorecard_status,
                   {_notes_sub}
            FROM interview i
            JOIN application  a  ON a.id  = i.application_id
            JOIN candidate    c  ON c.id  = a.candidate_id
            JOIN requisition  r  ON r.id  = a.requisition_id
            JOIN round_config rc ON rc.id = i.round_config_id
            WHERE {where}
            ORDER BY i.scheduled_at DESC NULLS LAST
            """,
            params,
        )

    return query(
        f"""
        SELECT i.id, i.scheduled_at, i.status, i.meet_link, i.mode,
               i.gcal_event_id,
               c.full_name AS candidate_name, r.title AS requisition,
               rc.name AS round_name, a.id AS application_id,
               {_panel_sub},
               {_sc_status_sub},
               {_notes_sub}
        FROM interview i
        JOIN application  a  ON a.id  = i.application_id
        JOIN candidate    c  ON c.id  = a.candidate_id
        JOIN requisition  r  ON r.id  = a.requisition_id
        JOIN round_config rc ON rc.id = i.round_config_id
        ORDER BY i.scheduled_at DESC NULLS LAST
        """,
        [uid, uid],
    )


# ─── Hiring-manager review ────────────────────────────────────────────────────

@router.get("/profiles-to-review")
def profiles_to_review(
    limit: int = Query(200, ge=1, le=500),
    offset: int = Query(0, ge=0),
    user: dict = Depends(get_current_user),
):
    uid = user["sub"]
    return query(
        """
        SELECT a.id, c.full_name, c.email, c.gender,
               r.title AS req_title, r.id AS req_id,
               a.combined_score, a.match_score, a.status,
               a.hm_feedback, a.hm_reviewed_at
        FROM application a
        JOIN candidate  c ON c.id  = a.candidate_id
        JOIN requisition r ON r.id = a.requisition_id
        WHERE r.hiring_manager_id = %s
          AND a.status = 'interview'
          AND (a.hm_feedback IS NULL OR a.hm_feedback = '')
        ORDER BY a.combined_score DESC NULLS LAST
        LIMIT %s OFFSET %s
        """,
        [uid, limit, offset],
    )


class HMFeedbackIn(BaseModel):
    approved: bool
    comment: Optional[str] = None


# ─── CV Database (Admin / TA Manager / Recruiter) ────────────────────────────

@router.get("/cv-database")
def cv_database(
    user: dict = Depends(get_current_user),
    limit: int = Query(200, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    role = user["role"]
    uid  = user["sub"]

    if role not in ("admin", "ta_manager", "recruiter"):
        raise HTTPException(403, "Not authorised to view CV database")

    # For recruiter, only show candidates from their requisitions
    if role == "recruiter":
        scope_where = """
            AND c.id IN (
                SELECT DISTINCT a_s.candidate_id
                FROM application a_s
                JOIN requisition_recruiter rr_s
                     ON rr_s.requisition_id = a_s.requisition_id
                     AND rr_s.recruiter_id = %s
            )
        """
        scope_params = [uid]
    else:
        scope_where  = ""
        scope_params = []

    summary = query_one(
        f"""
        SELECT
          COUNT(DISTINCT LOWER(c.email))                                    AS total_candidates,
          COUNT(a.id)                                                       AS total_applications,
          COUNT(DISTINCT LOWER(c.email)) FILTER
            (WHERE c.resume_url IS NOT NULL AND c.resume_url <> '')         AS resumes_on_file,
          ROUND(AVG(a.combined_score)
            FILTER (WHERE a.combined_score IS NOT NULL)::numeric, 1)        AS avg_score,
          COUNT(DISTINCT LOWER(c.email)) FILTER (WHERE a.status = 'hired') AS total_joined
        FROM candidate c
        LEFT JOIN application a ON a.candidate_id = c.id
        WHERE 1=1 {scope_where}
        """,
        scope_params,
    )

    # DISTINCT ON (email) collapses duplicate candidate records for the same person.
    # Keeps the oldest record (first created) as the canonical row.
    candidates = query(
        f"""
        SELECT * FROM (
          SELECT DISTINCT ON (LOWER(c.email))
            c.id, c.full_name, c.email, c.gender, c.source,
            c.resume_url,
            c.created_at                                                     AS registered_at,
            (SELECT COUNT(DISTINCT a_cnt.requisition_id)
             FROM application a_cnt
             JOIN candidate c_dup ON c_dup.id = a_cnt.candidate_id
             WHERE LOWER(c_dup.email) = LOWER(c.email))                     AS total_applications,
            (SELECT r.title
             FROM application a2
             JOIN candidate c2d ON c2d.id = a2.candidate_id
             JOIN requisition r ON r.id = a2.requisition_id
             WHERE LOWER(c2d.email) = LOWER(c.email)
             ORDER BY a2.applied_at DESC LIMIT 1)                           AS latest_position,
            (SELECT a3.status
             FROM application a3
             JOIN candidate c3d ON c3d.id = a3.candidate_id
             WHERE LOWER(c3d.email) = LOWER(c.email)
             ORDER BY a3.applied_at DESC LIMIT 1)                           AS latest_status,
            (SELECT a4.combined_score
             FROM application a4
             JOIN candidate c4d ON c4d.id = a4.candidate_id
             WHERE LOWER(c4d.email) = LOWER(c.email)
             ORDER BY a4.combined_score DESC NULLS LAST LIMIT 1)            AS best_score,
            (SELECT a5.bot_score
             FROM application a5
             JOIN candidate c5d ON c5d.id = a5.candidate_id
             WHERE LOWER(c5d.email) = LOWER(c.email)
             ORDER BY a5.bot_score DESC NULLS LAST LIMIT 1)                 AS ai_score,
            (SELECT a6.match_score
             FROM application a6
             JOIN candidate c6d ON c6d.id = a6.candidate_id
             WHERE LOWER(c6d.email) = LOWER(c.email)
             ORDER BY a6.match_score DESC NULLS LAST LIMIT 1)               AS match_score
          FROM candidate c
          WHERE 1=1 {scope_where}
          ORDER BY LOWER(c.email), c.created_at ASC
        ) deduped
        ORDER BY registered_at DESC
        LIMIT %s OFFSET %s
        """,
        scope_params + [limit, offset],
    )

    return {"summary": dict(summary) if summary else {}, "candidates": candidates}


@router.post("/applications/{app_id}/hm-feedback")
def hm_feedback(
    app_id: str,
    body: HMFeedbackIn,
    user: dict = Depends(get_current_user),
):
    role = user["role"]
    if role not in ("hiring_manager", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised to submit hiring-manager feedback")

    # Hiring manager may only give feedback on applications for their own requisitions
    if role == "hiring_manager":
        app_row = query_one(
            "SELECT requisition_id FROM application WHERE id=%s", [app_id]
        )
        req_id_str = str(app_row["requisition_id"]) if app_row and app_row["requisition_id"] else None
        if not req_id_str or not query_one(
            "SELECT 1 FROM requisition WHERE id=%s AND hiring_manager_id=%s",
            [req_id_str, user["sub"]],
        ):
            raise HTTPException(403, "Not authorised to submit feedback on this application")

    verdict = body.comment or ("Approved" if body.approved else "Not approved")
    row = query_one(
        """UPDATE application
           SET hm_feedback = %s, hm_reviewed_at = now()
           WHERE id = %s
           RETURNING id, status, hm_feedback""",
        [verdict, app_id],
    )
    if not row:
        raise HTTPException(404, "application not found")

    _maybe_send_feedback_package(app_id, verdict)
    return row


def _maybe_send_feedback_package(app_id: str, feedback_text: str) -> None:
    """This is a human panel interview — there's no AI transcript. If the
    interview was booked through the self-scheduling flow, package the HM's
    written feedback + candidate CV to the requisition's recruiters instead."""
    scheduled = query_one(
        """SELECT id FROM interview_schedule_request
           WHERE application_id = %s AND status = 'confirmed'
           ORDER BY confirmed_at DESC LIMIT 1""",
        [app_id],
    )
    if not scheduled:
        return

    ctx = query_one(
        """SELECT a.candidate_id, c.full_name AS candidate_name,
                  r.title AS job_title, r.id AS requisition_id, r.tenant_id AS tenant_id
           FROM application a
           JOIN candidate c ON c.id = a.candidate_id
           JOIN requisition r ON r.id = a.requisition_id
           WHERE a.id = %s""",
        [app_id],
    )
    if not ctx:
        return

    recruiter_rows = query(
        """SELECT au.email FROM requisition_recruiter rr
           JOIN app_user au ON au.id = rr.recruiter_id
           WHERE rr.requisition_id = %s AND au.is_active = TRUE""",
        [ctx["requisition_id"]],
    )
    recruiter_emails = [r["email"] for r in (recruiter_rows or []) if r.get("email")]
    if not recruiter_emails:
        return

    cv = connectors.load_candidate_cv_attachment(ctx["candidate_id"])
    attachments = [cv] if cv else []
    # TODO: attach the interview transcript here if panel interviews are ever recorded.
    subject = f"Panel interview feedback — {ctx['candidate_name']} — {ctx['job_title']}"
    body_txt = (
        f"Hiring-manager feedback for {ctx['candidate_name']} ({ctx['job_title']}):\n\n"
        f"{feedback_text}\n\n"
        f"{'Candidate CV attached.' if attachments else 'No CV on file for this candidate.'}\n\n"
        f"— Enternly"
    )
    from ..services.email_layout import build_branded_email
    html_body = build_branded_email(
        eyebrow="Application Tracking System",
        hero_title_html="Panel Interview<br>Feedback.",
        hero_subtitle="Written feedback from the hiring manager is ready for your review.",
        detail_cells=[("Candidate", ctx["candidate_name"]), ("Position", ctx["job_title"])],
        about_text=feedback_text + "\n\n" + (
            "Candidate CV attached." if attachments else "No CV on file for this candidate."
        ),
        about_heading="Feedback",
        cta_label=None, cta_link=None,
    )
    for addr in recruiter_emails:
        try:
            connectors.send_email(addr, subject, body_txt, html=html_body, attachments=attachments, tenant_id=ctx.get("tenant_id"))
        except Exception as exc:
            print(f"[pipeline] Failed to send HM feedback package to {addr}: {exc}")


# ─── Pipeline bifurcated view ─────────────────────────────────────────────────

@router.get("/requisitions/{req_id}/pipeline")
def get_req_pipeline(req_id: str, user: dict = Depends(get_current_user)):
    """
    Return all candidates for a requisition grouped by pipeline stage,
    with SLA/RAG status per candidate. Role-scoped.
    """
    _deny_hrbp(user)
    role, uid = user["role"], user["sub"]

    # Role scope: recruiter must own the req
    if role == "recruiter":
        if not query_one(
            "SELECT 1 FROM requisition_recruiter WHERE requisition_id=%s AND recruiter_id=%s",
            [req_id, uid],
        ):
            raise HTTPException(403, "Not authorised")

    rows = query(
        """
        SELECT a.id AS app_id, a.status, a.ai_fit_score, a.bot_score,
               a.combined_score, a.match_score, a.current_round,
               a.screening_decision,
               c.full_name, c.email,
               iv.id AS interview_id, iv.scheduled_at AS interview_scheduled_at, iv.status AS interview_status,
               EXTRACT(EPOCH FROM (
                   now() - COALESCE(
                       (SELECT se.occurred_at FROM stage_event se
                        WHERE se.application_id=a.id AND se.to_status=a.status
                        ORDER BY se.occurred_at DESC LIMIT 1),
                       a.applied_at
                   )
               ))/86400.0 AS elapsed_days
        FROM application a
        JOIN candidate c ON c.id=a.candidate_id
        LEFT JOIN LATERAL (
            SELECT i.id, i.scheduled_at, i.status
            FROM interview i
            JOIN round_config rc ON rc.id = i.round_config_id
            WHERE i.application_id = a.id AND rc.sequence = a.current_round
            ORDER BY i.created_at DESC LIMIT 1
        ) iv ON true
        WHERE a.requisition_id=%s
        ORDER BY a.applied_at DESC
        """,
        [req_id],
    ) or []

    cfg = load_config(user.get("tenant_id"))

    # Fetch round_config for dynamic interview level names
    round_cfg = query(
        "SELECT id, sequence, name, round_type, panelist_emails, is_auto FROM round_config WHERE requisition_id=%s ORDER BY sequence",
        [req_id],
    ) or []
    round_names     = {r["sequence"]: r["name"] for r in round_cfg}
    round_ids       = {r["sequence"]: r["id"] for r in round_cfg}
    round_types     = {r["sequence"]: r["round_type"] for r in round_cfg}
    round_panelists = {r["sequence"]: list(r["panelist_emails"] or []) for r in round_cfg}
    round_is_auto   = {r["sequence"]: r["is_auto"] for r in round_cfg}
    max_interview_rounds = max((r["sequence"] for r in round_cfg), default=0)

    stage_map = {s: [] for s in PIPELINE_STAGES}
    terminals = {"hired": [], "rejected": [], "on_hold": []}

    for r in rows:
        status = r["status"]
        elapsed = float(r["elapsed_days"] or 0)
        if status in TERMINAL:
            rag = None
        else:
            sla_key = STAGE_SLA_KEY.get(status, "stage_default")
            tgt = cfg.get(sla_key, cfg.get("stage_default", 5))
            rag = compute_rag(elapsed, tgt)

        entry = {
            "app_id":            str(r["app_id"]),
            "full_name":         r["full_name"],
            "email":             r["email"],
            "elapsed_days":      round(elapsed, 1),
            "ai_fit_score":      r["ai_fit_score"],
            "bot_score":         r["bot_score"],
            "combined_score":    r["combined_score"],
            "current_round":     r["current_round"],
            "screening_decision": r["screening_decision"],
            "rag":               rag,
            "interview_id":           str(r["interview_id"]) if r["interview_id"] else None,
            "interview_scheduled_at": r["interview_scheduled_at"].isoformat() if r["interview_scheduled_at"] else None,
            "interview_status":       r["interview_status"],
        }
        if status in stage_map:
            stage_map[status].append(entry)
        elif status in terminals:
            terminals[status].append(entry)

    # Build per-level sub-groups for the interview stage
    interview_candidates = stage_map["interview"]
    level_map: dict[int, list] = {}
    for c in interview_candidates:
        lvl = int(c["current_round"] or 1)
        level_map.setdefault(lvl, []).append(c)
    interview_levels = [
        {
            "level":            lvl,
            "round_name":       round_names.get(lvl, f"Level {lvl}"),
            "round_config_id":  round_ids.get(lvl),
            "round_type":       round_types.get(lvl),
            "is_auto":          bool(round_is_auto.get(lvl, False)),
            "panelist_emails":  round_panelists.get(lvl, []),
            "count":            len(cands),
            "candidates":       cands,
        }
        for lvl, cands in sorted(level_map.items())
    ]

    req = query_one("SELECT title, req_code FROM requisition WHERE id=%s", [req_id])

    stages_out = []
    for s in PIPELINE_STAGES:
        stage_entry = {
            "stage":      s,
            "label":      PIPELINE_STAGE_LABELS.get(s, s.replace("_", " ").title()),
            "count":      len(stage_map[s]),
            "candidates": stage_map[s],
        }
        if s == "interview":
            stage_entry["levels"] = interview_levels
        stages_out.append(stage_entry)

    return {
        "req_id":               req_id,
        "title":                req["title"] if req else "",
        "req_code":             req["req_code"] if req else "",
        "max_interview_rounds": max_interview_rounds,
        "round_config":         [{"sequence": r["sequence"], "name": r["name"], "id": r["id"], "round_type": r["round_type"], "panelist_emails": list(r["panelist_emails"] or [])} for r in round_cfg],
        "stages":               stages_out,
        "terminal":             {k: {"count": len(v), "candidates": v} for k, v in terminals.items()},
    }


# ─── One-click advance ────────────────────────────────────────────────────────

class AdvanceIn(BaseModel):
    target: Optional[str] = None   # "rejected" | "on_hold" | "hired" for terminal; None = auto-next


@router.post("/applications/{app_id}/advance")
def advance_application(
    app_id: str,
    body: AdvanceIn = AdvanceIn(),
    user: dict = Depends(get_current_user),
):
    """Advance (or terminate) a candidate in the pipeline. Logs to stage_event."""
    role = user["role"]
    if role not in ("recruiter", "ta_manager", "admin", "hiring_manager"):
        raise HTTPException(403, "Not authorised to advance candidates")

    app = query_one(
        "SELECT id, status, requisition_id, current_round, applied_at FROM application WHERE id=%s", [app_id]
    )
    if not app:
        raise HTTPException(404, "Application not found")

    # Hiring manager may only advance on their own requisitions
    if role == "hiring_manager":
        req_id_str = str(app["requisition_id"]) if app["requisition_id"] else None
        if not req_id_str or not query_one(
            "SELECT 1 FROM requisition WHERE id=%s AND hiring_manager_id=%s",
            [req_id_str, user["sub"]],
        ):
            raise HTTPException(403, "Not authorised to advance candidates on this requisition")

    # Recruiter may only advance candidates on requisitions they own
    if role == "recruiter":
        if not _recruiter_owns_req(user, str(app["requisition_id"]) if app["requisition_id"] else None):
            raise HTTPException(404, "Application not found")

    current = app["status"]
    req_id  = str(app["requisition_id"]) if app["requisition_id"] else None

    # Terminal move (reject / hold / hire)
    if body.target in ("rejected", "on_hold", "hired"):
        if current in TERMINAL:
            raise HTTPException(400, f"Already in terminal status '{current}'")
        with transaction() as cur:
            tx_exec(cur, "UPDATE application SET status=%s WHERE id=%s", [body.target, app_id])
            tx_exec(
                cur,
                "INSERT INTO stage_event (application_id, from_status, to_status, actor_id) VALUES (%s,%s,%s,%s)",
                [app_id, current, body.target, user["sub"]],
            )
        sync_plan_on_advance(app_id, body.target, current, req_id)
        if body.target == "rejected":
            _send_application_rejected_email(app_id, user)
        # Gamification: quality event when candidate joins; volume for every advance
        if body.target == "hired":
            _gam_award("recruiter", user["sub"], "offer_joined", req_id, app_id)
        _gam_award("recruiter", user["sub"], "candidate_advanced", req_id, app_id)
        return {"ok": True, "prev_stage": current, "new_stage": body.target}

    # Auto advance
    if current in TERMINAL:
        raise HTTPException(400, f"Application is in terminal status '{current}'")

    # ── Dynamic interview level advance ───────────────────────────────────────
    if current == "interview":
        current_round = int(app["current_round"] or 1)
        max_row = query_one(
            "SELECT COALESCE(MAX(sequence),0) AS max_seq FROM round_config WHERE requisition_id=%s",
            [req_id],
        )
        max_seq = int((max_row or {}).get("max_seq") or 1)

        if current_round < max_seq:
            # Advance to next interview level (status stays 'interview')
            new_round = current_round + 1
            round_name_row = query_one(
                "SELECT name FROM round_config WHERE requisition_id=%s AND sequence=%s",
                [req_id, new_round],
            )
            round_label = (round_name_row or {}).get("name") or f"Level {new_round}"
            with transaction() as cur:
                tx_exec(cur, "UPDATE application SET current_round=%s WHERE id=%s", [new_round, app_id])
                tx_exec(
                    cur,
                    "INSERT INTO stage_event (application_id, from_status, to_status, actor_id, note) VALUES (%s,%s,%s,%s,%s)",
                    [app_id, current, current, user["sub"], f"Interview level {new_round}: {round_label}"],
                )
                sched_row, sched_ctx = _schedule_request_for_round_tx(cur, req_id, app_id, new_round, user["sub"])
            hm_notified = _schedule_side_effects(sched_row, sched_ctx, app_id, user["sub"])
            _gam_award("recruiter", user["sub"], "candidate_advanced", req_id, app_id)
            return {
                "ok":         True,
                "prev_stage": "interview",
                "new_stage":  "interview",
                "level":      new_round,
                "level_name": round_label,
                "hm_notified": hm_notified,
            }
        else:
            # All levels complete — advance to documentation
            with transaction() as cur:
                tx_exec(cur, "UPDATE application SET status='documentation' WHERE id=%s", [app_id])
                tx_exec(
                    cur,
                    "INSERT INTO stage_event (application_id, from_status, to_status, actor_id) VALUES (%s,%s,%s,%s)",
                    [app_id, "interview", "documentation", user["sub"]],
                )
            _gam_award("recruiter", user["sub"], "candidate_advanced", req_id, app_id)
            return {
                "ok":         True,
                "prev_stage": "interview",
                "new_stage":  "documentation",
                "needs_offer": True,
            }

    # ── Standard next-stage advance ──────────────────────────────────────────
    next_stage = NEXT_STAGE.get(current)
    if not next_stage:
        raise HTTPException(400, f"No next stage after '{current}' (already at end of pipeline)")

    extra_set = ""
    extra_params: list = [app_id]
    # Skip NexAI if the requisition has no bot_interview round configured —
    # go straight to shortlisted. Mirrors the interview-skip logic below.
    if next_stage == "nexai_bot":
        bot_round = query_one(
            "SELECT 1 FROM round_config WHERE requisition_id=%s AND round_type='bot_interview' LIMIT 1",
            [req_id],
        )
        if not bot_round:
            next_stage = "shortlisted"

    # Enter the first PANEL round when transitioning into interview for the
    # first time. round_config.sequence numbers a requisition's rounds
    # end-to-end (bot_interview included), and a bot_interview round always
    # occupies sequence 1 when configured -- so the first *panel* round is
    # not necessarily sequence 1. current_round must equal the real
    # round_config.sequence value (the kanban board matches candidates into
    # interview columns by c.current_round === round.sequence, and
    # _schedule_request_for_round_tx looks panelists up the same way), so it
    # has to be resolved here rather than hardcoded.
    first_panel_seq = None
    if next_stage == "interview":
        # Skip interview if no panel rounds are configured — go straight to documentation
        first_panel_row = query_one(
            "SELECT MIN(sequence) AS seq FROM round_config WHERE requisition_id=%s AND round_type<>'bot_interview'",
            [req_id],
        )
        first_panel_seq = (first_panel_row or {}).get("seq")
        if first_panel_seq is None:
            next_stage = "documentation"
        else:
            extra_set = ", current_round=%s"
            extra_params = [first_panel_seq, app_id]

    with transaction() as cur:
        tx_exec(
            cur,
            f"UPDATE application SET status=%s{extra_set} WHERE id=%s",
            [next_stage] + extra_params,
        )
        tx_exec(
            cur,
            "INSERT INTO stage_event (application_id, from_status, to_status, actor_id) VALUES (%s,%s,%s,%s)",
            [app_id, current, next_stage, user["sub"]],
        )
        if next_stage == "interview" and extra_set:
            sched_row, sched_ctx = _schedule_request_for_round_tx(cur, req_id, app_id, first_panel_seq, user["sub"])

    if next_stage == "interview" and extra_set:
        flags_hm_notified = _schedule_side_effects(sched_row, sched_ctx, app_id, user["sub"])
    else:
        flags_hm_notified = None

    flags = {}
    if flags_hm_notified is not None:
        flags["hm_notified"] = flags_hm_notified
    if next_stage == "nexai_bot":
        flags["needs_nexai_invite"] = True
    elif next_stage == "documentation":
        flags["needs_offer"] = True

    sync_plan_on_advance(app_id, next_stage, current, req_id)

    # Gamification: volume point for every pipeline advance
    _gam_award("recruiter", user["sub"], "candidate_advanced", req_id, app_id)
    # Speed bonus: screen completed within 24 h of application submission
    if current == "applied":
        applied_at = app.get("applied_at")
        if applied_at:
            from datetime import datetime, timezone
            _now = datetime.now(timezone.utc)
            _aa = applied_at if applied_at.tzinfo else applied_at.replace(tzinfo=timezone.utc)
            if (_now - _aa).total_seconds() < 86400:
                _gam_award("recruiter", user["sub"], "fast_screen", req_id, app_id)

    return {"ok": True, "prev_stage": current, "new_stage": next_stage, **flags}


# ─── Admin/TA-manager: correct a mistakenly-advanced candidate ───────────────
# The state machine above only ever moves forward -- there is deliberately no
# "undo" in the normal flow, since going backward isn't a real hiring action.
# This is a narrow escape hatch for the "moved to the wrong stage by mistake"
# case: admin/ta_manager only, one candidate at a time, always leaves an
# explicit stage_event + activity_log entry so it's fully auditable, and
# cancels (never deletes) any in-flight offer it would otherwise orphan.

class RevertStageIn(BaseModel):
    to_stage: str
    round_index: Optional[int] = None   # required when to_stage == 'interview'
    note: Optional[str] = None


def _revert_cur_idx(app_id: str, current: str) -> int:
    """Index (exclusive upper bound) into PIPELINE_STAGES of stages this
    application can be reverted to. Shared by the options-lookup GET and the
    revert POST so the two can never drift apart."""
    if current == "rejected":
        # A rejection can land from any pipeline stage -- look up the stage
        # the candidate was actually rejected from (stage_event.from_status)
        # so the modal only offers that stage and earlier, rather than
        # guessing from a fixed order.
        pre_reject = query_one(
            """SELECT from_status FROM stage_event
               WHERE application_id=%s AND to_status='rejected'
               ORDER BY occurred_at DESC LIMIT 1""",
            [app_id],
        )
        pre_stage = (pre_reject or {}).get("from_status")
        if pre_stage not in PIPELINE_STAGES:
            raise HTTPException(400, "Cannot determine which stage this candidate was rejected from")
        return PIPELINE_STAGES.index(pre_stage) + 1
    if current in TERMINAL:
        raise HTTPException(400, f"Application is in terminal status '{current}' — reopen it first, this isn't a stage revert")
    if current in PIPELINE_STAGES:
        return PIPELINE_STAGES.index(current)
    raise HTTPException(400, f"Current status '{current}' is not a revertible pipeline stage")


@router.get("/applications/{app_id}/revert-options")
def get_revert_options(app_id: str, user: dict = Depends(get_current_user)):
    if user["role"] not in ("admin", "ta_manager"):
        raise HTTPException(403, "Only an admin or TA manager can revert a candidate's stage")
    app = query_one("SELECT id, status FROM application WHERE id=%s", [app_id])
    if not app:
        raise HTTPException(404, "Application not found")
    try:
        cur_idx = _revert_cur_idx(app_id, app["status"])
    except HTTPException:
        # Nothing earlier to offer (e.g. still at 'applied', or a terminal
        # status revert doesn't apply to) -- an empty option list, not an error.
        return {"options": []}
    return {"options": PIPELINE_STAGES[:cur_idx]}


@router.post("/applications/{app_id}/revert-stage")
def revert_application_stage(
    app_id: str,
    body: RevertStageIn,
    user: dict = Depends(get_current_user),
):
    if user["role"] not in ("admin", "ta_manager"):
        raise HTTPException(403, "Only an admin or TA manager can revert a candidate's stage")

    app = query_one(
        "SELECT id, status, requisition_id, current_round FROM application WHERE id=%s", [app_id]
    )
    if not app:
        raise HTTPException(404, "Application not found")
    current = app["status"]
    req_id = str(app["requisition_id"]) if app["requisition_id"] else None

    cur_idx = _revert_cur_idx(app_id, current)
    if body.to_stage not in PIPELINE_STAGES:
        raise HTTPException(400, f"to_stage must be one of {PIPELINE_STAGES}")
    target_idx = PIPELINE_STAGES.index(body.to_stage)
    if target_idx >= cur_idx:
        raise HTTPException(400, f"'{body.to_stage}' is not earlier than the current stage '{current}'")

    extra_set, extra_params = "", []
    if body.to_stage == "interview":
        if not body.round_index:
            raise HTTPException(400, "round_index is required when reverting to 'interview'")
        rc = query_one(
            "SELECT 1 FROM round_config WHERE requisition_id=%s AND sequence=%s",
            [req_id, body.round_index],
        )
        if not rc:
            raise HTTPException(400, f"No round configured at sequence {body.round_index} for this requisition")
        extra_set = ", current_round=%s"
        extra_params = [body.round_index]

    note = body.note or f"Reverted by {user['role']} from '{current}' to '{body.to_stage}'"

    with transaction() as cur:
        tx_exec(
            cur,
            f"UPDATE application SET status=%s{extra_set} WHERE id=%s",
            [body.to_stage] + extra_params + [app_id],
        )
        tx_exec(
            cur,
            "INSERT INTO stage_event (application_id, from_status, to_status, actor_id, note) VALUES (%s,%s,%s,%s,%s)",
            [app_id, current, body.to_stage, user["sub"], note],
        )
        # An offer created for this application (documentation/offered stages)
        # would otherwise be silently orphaned -- cancel it (never delete; it
        # stays in the audit trail) rather than leave a 'pending_approval'/
        # 'approved' offer floating for an application that's back in interview.
        offer_cancelled = False
        existing_offer = tx_exec(
            cur, "SELECT id, status FROM offer WHERE application_id=%s", [app_id]
        )
        if existing_offer and existing_offer[0]["status"] not in ("cancelled", "rejected", "declined"):
            tx_exec(
                cur, "UPDATE offer SET status='cancelled', updated_at=now() WHERE id=%s",
                [existing_offer[0]["id"]],
            )
            offer_cancelled = True

    log_activity(
        "application", "stage_reverted",
        entity_id=app_id, application_id=app_id, requisition_id=req_id,
        actor_id=user["sub"], actor_role=user["role"],
        detail={
            "from_status": current, "to_status": body.to_stage,
            "round_index": body.round_index, "note": note,
            "offer_cancelled": offer_cancelled,
        },
    )
    return {
        "ok": True, "prev_stage": current, "new_stage": body.to_stage,
        "offer_cancelled": offer_cancelled,
    }


# ─── Manual screening decision ───────────────────────────────────────────────

class ScreenDecisionIn(BaseModel):
    decision: str   # "pass" | "hold" | "reject"
    notes: Optional[str] = None


@router.post("/applications/{app_id}/screen-decision")
def record_screen_decision(
    app_id: str,
    body: ScreenDecisionIn,
    user: dict = Depends(get_current_user),
):
    """Record recruiter's manual screening decision (pass/hold/reject) on an application."""
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")
    if user["role"] == "recruiter":
        req_id = _application_req_id(app_id)
        if not req_id or not _recruiter_owns_req(user, req_id):
            raise HTTPException(404, "Application not found")
    if body.decision not in ("pass", "hold", "reject"):
        raise HTTPException(400, "decision must be 'pass', 'hold', or 'reject'")

    row = query_one(
        """UPDATE application
           SET screening_decision=%s, screening_notes=%s,
               screened_by=%s, screened_at=now()
           WHERE id=%s
           RETURNING id, status, screening_decision""",
        [body.decision, body.notes, user["sub"], app_id],
    )
    if not row:
        raise HTTPException(404, "Application not found")

    # Reject decision immediately moves to terminal
    if body.decision == "reject":
        query("UPDATE application SET status='rejected' WHERE id=%s", [app_id], fetch=False)
        query(
            "INSERT INTO stage_event (application_id, from_status, to_status, actor_id, note) VALUES (%s,'screen','rejected',%s,'Rejected at screening')",
            [app_id, user["sub"]], fetch=False,
        )
        _send_application_rejected_email(app_id, user)
    else:
        # pass/hold write no stage_event today — activity_log is the only record
        log_activity(
            "application", f"screen_decision_{body.decision}",
            application_id=app_id, requisition_id=_application_req_id(app_id),
            actor_id=user["sub"], actor_role=user["role"],
            to_value=body.decision, detail={"notes": body.notes},
        )

    return row


# ─── Delete application (fix a wrongly-added candidate) ──────────────────────

def _delete_application_row(app_id: str) -> tuple[bool, Optional[str]]:
    """
    Shared core for single + bulk application delete. Returns (ok, reason).
    `reason` is a human-readable block/not-found message when ok is False.

    Refuses once an offer or a proctoring/interview-integrity record exists
    for the application — those are compliance records and must go through
    the normal reject/withdraw flow instead of a hard delete. interview,
    stage_event, nexai_invite/session, application_document and
    negotiation_log rows cascade automatically via their table definitions;
    candidate_feedback and the gamification ledger just lose the link
    (ON DELETE SET NULL) so those records survive.
    """
    app_row = query_one("SELECT id, candidate_id FROM application WHERE id=%s", [app_id])
    if not app_row:
        return False, "Application not found"

    if query_one("SELECT id FROM offer WHERE application_id=%s LIMIT 1", [app_id]):
        return False, "An offer exists for this application — reject the candidate instead"
    if query_one("SELECT id FROM proctoring_session WHERE application_id=%s LIMIT 1", [app_id]):
        return False, "A proctoring/interview record exists for this application — reject the candidate instead"
    if query_one("SELECT id FROM proctoring_appeal WHERE application_id=%s LIMIT 1", [app_id]):
        return False, "A proctoring appeal exists for this application — reject the candidate instead"

    # Detach optional/historical references so they don't block the delete via FK.
    query("UPDATE campus_candidate SET application_id=NULL, nexai_session_id=NULL WHERE application_id=%s", [app_id], fetch=False)
    query("UPDATE sent_email_log SET application_id=NULL WHERE application_id=%s", [app_id], fetch=False)

    query("DELETE FROM application WHERE id=%s", [app_id], fetch=False)

    # Best-effort: a candidate left with zero applications is an orphan record — clean it up too.
    remaining = query_one("SELECT COUNT(*) AS n FROM application WHERE candidate_id=%s", [app_row["candidate_id"]])
    if remaining and remaining["n"] == 0:
        try:
            query("DELETE FROM candidate WHERE id=%s", [app_row["candidate_id"]], fetch=False)
        except Exception:
            pass

    return True, None


@router.delete("/applications/{app_id}")
def delete_application(app_id: str, user: dict = Depends(get_current_user)):
    """Permanently remove a single wrongly-added candidate application. See _delete_application_row."""
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")
    if user["role"] == "recruiter":
        req_id = _application_req_id(app_id)
        if not req_id or not _recruiter_owns_req(user, req_id):
            raise HTTPException(404, "Application not found")

    ok, reason = _delete_application_row(app_id)
    if not ok:
        raise HTTPException(404 if reason == "Application not found" else 400, reason)
    return {"ok": True}


class BulkDeleteApplicationsIn(BaseModel):
    app_ids: list[str]


@router.post("/applications/bulk-delete")
def bulk_delete_applications(body: BulkDeleteApplicationsIn, user: dict = Depends(get_current_user)):
    """
    Delete multiple wrongly-added candidate applications at once (e.g. a bad
    bulk campus upload). Each id is evaluated independently against the same
    offer/proctoring safety checks as the single-delete endpoint — one
    blocked row never stops the rest of the batch from being removed.
    """
    if user["role"] not in ("recruiter", "ta_manager", "admin"):
        raise HTTPException(403, "Not authorised")
    if not body.app_ids:
        raise HTTPException(400, "app_ids must be a non-empty list")

    deleted: list[str] = []
    failed: list[dict] = []
    for app_id in body.app_ids:
        if user["role"] == "recruiter":
            req_id = _application_req_id(app_id)
            if not req_id or not _recruiter_owns_req(user, req_id):
                failed.append({"app_id": app_id, "reason": "Not found"})
                continue
        ok, reason = _delete_application_row(app_id)
        if ok:
            deleted.append(app_id)
        else:
            failed.append({"app_id": app_id, "reason": reason})

    return {"deleted": len(deleted), "failed": failed}


# ─── Per-stage report ─────────────────────────────────────────────────────────

@router.get("/requisitions/{req_id}/stage/{stage}/report")
def stage_report(req_id: str, stage: str, user: dict = Depends(get_current_user)):
    """Return all candidates in a given stage for CSV export. Role-scoped."""
    role, uid = user["role"], user["sub"]
    if role == "recruiter":
        if not query_one(
            "SELECT 1 FROM requisition_recruiter WHERE requisition_id=%s AND recruiter_id=%s",
            [req_id, uid],
        ):
            raise HTTPException(403, "Not authorised")

    rows = query(
        """
        SELECT a.id AS app_id, c.full_name, c.email, c.gender,
               a.status, a.ai_fit_score, a.bot_score, a.combined_score,
               a.current_company, a.current_designation,
               a.notice_period_days, a.current_ctc_total, a.expected_ctc_total,
               a.applied_at,
               EXTRACT(EPOCH FROM (
                   now() - COALESCE(
                       (SELECT se.occurred_at FROM stage_event se
                        WHERE se.application_id=a.id AND se.to_status=a.status
                        ORDER BY se.occurred_at DESC LIMIT 1),
                       a.applied_at
                   )
               ))/86400.0 AS days_in_stage
        FROM application a
        JOIN candidate c ON c.id=a.candidate_id
        WHERE a.requisition_id=%s AND a.status=%s
        ORDER BY a.combined_score DESC NULLS LAST
        """,
        [req_id, stage],
    ) or []

    req = query_one("SELECT title, req_code FROM requisition WHERE id=%s", [req_id])

    return {
        "requisition": dict(req) if req else {},
        "stage": stage,
        "stage_label": PIPELINE_STAGE_LABELS.get(stage, stage),
        "count": len(rows),
        "rows": [dict(r) for r in rows],
    }


# ── Requisition criticality (TA admin) ───────────────────────────────────────

_CRITICALITY_VALUES = {"Low", "Medium", "High", "Critical"}


class CriticalityIn(BaseModel):
    criticality: str


@router.patch("/requisitions/{req_id}/criticality")
def patch_criticality(req_id: str, body: CriticalityIn, user: dict = Depends(get_current_user)):
    """Set the criticality flag on a requisition. Controls gamification score multipliers."""
    if user.get("role") not in ("ta_manager", "admin"):
        raise HTTPException(403, "ta_manager / admin only")
    if body.criticality not in _CRITICALITY_VALUES:
        raise HTTPException(400, "criticality must be Low | Medium | High | Critical")
    if not query_one("SELECT id FROM requisition WHERE id=%s", [req_id]):
        raise HTTPException(404, "Requisition not found")
    query("UPDATE requisition SET criticality=%s WHERE id=%s", [body.criticality, req_id], fetch=False)
    return {"ok": True, "req_id": req_id, "criticality": body.criticality}
